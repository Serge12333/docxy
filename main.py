import tkinter
from tkinter import *
from tkinter import messagebox
from tkinter.ttk import *
from tkinter import ttk
import tkinter as tk
import os
from mailmerge import MailMerge
from transliterate import translit
import tkentrycomplete
import json
import sys
from datetime import datetime, timedelta, date
import shutil
import openpyxl
from openpyxl.styles import numbers
from decimal import Decimal
import uuid

# --- Globals ---
# Main container for dynamically created widgets
dynamic_frame = None
# Global dictionary to hold the state of dynamically created checkboxes
checkbox_vars = {}
# Global dictionary to hold selections from main-key comboboxes
main_key_selections = {}

# Determine the base directory
if getattr(sys, 'frozen', False):
    # If the application is frozen (i.e., deployed with PyInstaller),
    # use the temporary path where files are extracted.
    BASE_DIR = sys._MEIPASS
else:
    # If the application is run as a regular Python script,
    # use the directory of the script file.
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# The JSON folder is inside the base directory
JSON_DIR = os.path.join(BASE_DIR, 'json')

# Construct the full paths to each JSON file inside the 'json' folder.
FIELDS_CONFIG_PATH = os.path.join(JSON_DIR, 'fields_config.json')
COMBOBOX_REGULAR_PATH = os.path.join(JSON_DIR, 'combobox_regular.json')
COMBOBOX_MAINKEY_PATH = os.path.join(JSON_DIR, 'combobox_mainkey.json')
COMBINATION_CONFIG_PATH = os.path.join(JSON_DIR, 'combination_config.json')
RULES_CONFIG_PATH = os.path.join(JSON_DIR, 'rules_config.json')
ALL_TAGS_OUTPUT_PATH = os.path.join(JSON_DIR, 'all_tags.json')


# --- Utility and Core Logic Functions ---

def _onKeyRelease(event):
    """Handles Ctrl+C, Ctrl+V, Ctrl+X for entry widgets."""
    ctrl = (event.state & 0x4) != 0
    if event.keycode == 88 and ctrl and event.keysym.lower() != "x":
        event.widget.event_generate("<<Cut>>")
    elif event.keycode == 86 and ctrl and event.keysym.lower() != "v":
        event.widget.event_generate("<<Paste>>")
    elif event.keycode == 67 and ctrl and event.keysym.lower() != "c":
        event.widget.event_generate("<<Copy>>")


def load_json(file_path, dict_name):
    """Loads data from a JSON file, creating it if it doesn't exist."""
    try:
        with open(file_path, 'r', encoding='utf8') as f:
            return json.load(f)
    except FileNotFoundError:
        with open(file_path, 'w', encoding='utf8') as f:
            json.dump([], f, ensure_ascii=False, indent=4)
        messagebox.showwarning("Информация", f"Создан новый файл конфигурации: {file_path}")
        return []
    except json.JSONDecodeError:
        messagebox.showwarning("Предупреждение", f"Файл {file_path} поврежден. Инициализация пустой конфигурации.")
        with open(file_path, 'w', encoding='utf8') as f:
            json.dump([], f, ensure_ascii=False, indent=4)
        return []


def save_json(file_path, data):
    """Safely writes data to a JSON file to prevent corruption."""
    temp_file = file_path + '.tmp'
    try:
        with open(temp_file, 'w', encoding='utf8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        os.replace(temp_file, file_path)
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить конфигурацию в {file_path}: {e}")
        if os.path.exists(temp_file):
            os.remove(temp_file)

def populate_tags_listbox(tags_listbox):
    """Populates the tags listbox in the constructor window."""
    for i in tags_listbox.get_children():
        tags_listbox.delete(i)

    all_items = []
    all_items.extend(
        [(f['name'], f['type'], f.get('tag_type', 'поле')) for f in load_json(FIELDS_CONFIG_PATH, 'fields_config')])
    all_items.extend([(c['name'], c['type'], c.get('tag_type', 'комбобокс')) for c in
                      load_json(COMBOBOX_REGULAR_PATH, 'combobox_regular')])
    all_items.extend([(c['name'], c['type'], c.get('tag_type', 'список')) for c in
                      load_json(COMBOBOX_MAINKEY_PATH, 'combobox_mainkey')])

    try:
        combination_config = load_combination_config()
        for combo in combination_config:
            all_items.append((combo['name'], combo['type'], combo.get('tag_type', 'сочетание')))
    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка при загрузке combination_config: {str(e)}")

    all_items.sort(key=lambda x: x[0])
    for item in all_items:
        tags_listbox.insert("", "end", values=item)


def get_current_widget_values(frame):
    """Saves the current values from widgets in the given frame."""
    state = {}
    if not frame:
        return state
    for widget in frame.winfo_children():
        if not hasattr(widget, '_name'):
            continue
        widget_name = widget._name
        if isinstance(widget, (tk.Entry, tkentrycomplete.Combobox)):
            state[widget_name] = widget.get()
        elif isinstance(widget, ttk.Checkbutton):
            var = checkbox_vars.get(widget_name)
            if var:
                state[widget_name] = var.get()  # Save the integer value (0 or 1)
    return state


def refresh_all_windows(listbox_to_refresh):
    """Refreshes the dynamic widgets in the main window and the constructor listbox."""
    global dynamic_frame

    # 1. Save current state before destroying widgets
    current_state = get_current_widget_values(dynamic_frame)

    if dynamic_frame and dynamic_frame.winfo_exists():
        for widget in dynamic_frame.winfo_children():
            widget.destroy()
        # 2. Pass the saved state to the loading function
        load_all_dynamic_widgets(initial_state=current_state)

    # Refresh the constructor's listbox
    if listbox_to_refresh and listbox_to_refresh.winfo_exists():
        populate_tags_listbox(listbox_to_refresh)


def export_all_tags_to_json(parent_window):
    """
    Silently collects all available tags, sorts them alphabetically,
    and saves them to a single JSON file. Only shows a message on error.
    """
    try:
        all_tags = set()

        # 1. Add the hardcoded 'today_tag'
        all_tags.add('today_tag')

        # 2. Load tags from fields_config.json
        fields_config = load_json(FIELDS_CONFIG_PATH, 'fields_config')
        for field in fields_config:
            all_tags.add(field['name'])

        # 3. Load tags from combobox_regular.json
        regular_combos = load_json(COMBOBOX_REGULAR_PATH, 'combobox_regular')
        for combo in regular_combos:
            all_tags.add(combo['name'])

        # 4. Load from combobox_mainkey.json (both main keys and all sub-keys)
        mainkey_combos = load_json(COMBOBOX_MAINKEY_PATH, 'combobox_mainkey')
        for combo in mainkey_combos:
            all_tags.add(combo['name'])  # Add the main key itself
            for mk_dict in combo.get('main_keys', []):
                if mk_dict:
                    subkeys_dict = list(mk_dict.values())[0]
                    for subkey_name in subkeys_dict.keys():
                        all_tags.add(subkey_name)

        # 5. Load from combination_config.json (only the combination's name)
        combination_config = load_json(COMBINATION_CONFIG_PATH, 'combination_config')
        for combo in combination_config:
            all_tags.add(combo['name'])

        # Sort the final list alphabetically (case-insensitive)
        sorted_tags_list = sorted(list(all_tags), key=str.lower)

        # Save the sorted list to the new JSON file
        save_json(ALL_TAGS_OUTPUT_PATH, sorted_tags_list)

        # Print a silent confirmation to the console instead of a popup
        print(f"Tag list updated: {len(sorted_tags_list)} tags exported to {ALL_TAGS_OUTPUT_PATH}")

    except Exception as e:
        # Still show an error if something goes wrong
        messagebox.showerror("Ошибка экспорта тегов", f"Произошла ошибка при автоматическом экспорте тегов: {e}",
                             parent=parent_window)

def on_constructor_close(window_to_destroy):
    """Handles the closing event for the constructor window."""
    # The main 'window' is passed as the parent for any potential error message
    export_all_tags_to_json(window)
    window_to_destroy.destroy()

def get_common_merge_data():
    """Collects data from all dynamically created UI elements for the mail merge."""
    global dynamic_frame
    merge_data = {}

    # Add the only hardcoded tag value
    merge_data['today_tag'] = datetime.now().strftime('%d.%m.%Y')
    try:
        all_tags_from_file = load_json(ALL_TAGS_OUTPUT_PATH, 'all_tags.json')
        if isinstance(all_tags_from_file, list):
            # Filter out any potential non-string items and join them
            string_tags = [str(tag) for tag in all_tags_from_file]
            merge_data['all_tags_merge'] = ', '.join(string_tags)
        else:
            # Failsafe in case the JSON is not a list
            merge_data['all_tags_merge'] = ''
    except Exception as e:
        print(f"Could not create 'all_tags_merge' due to an error: {e}")
        merge_data['all_tags_merge'] = ''

    # Collect data from dynamically created widgets
    if dynamic_frame:
        for widget in dynamic_frame.winfo_children():
            if not hasattr(widget, '_name'):
                continue

            widget_name = widget._name
            if isinstance(widget, (tk.Entry, tkentrycomplete.Combobox)):
                merge_data[widget_name] = widget.get()
            elif isinstance(widget, ttk.Checkbutton):
                var = checkbox_vars.get(widget_name)
                if var:
                    merge_data[widget_name] = "1" if var.get() else "0"

    # Include subkeys from main_key combobox selections
    for data_dict in main_key_selections.values():
        merge_data.update(data_dict)

    # Include combined tags
    combination_config = load_combination_config()  # Make sure this function exists
    for combo in combination_config:
        combined_value = ""
        for tag in combo['tags']:
            # All tags, including 'today_tag', can now be looked up directly
            if tag == 'today_tag':
                combined_value += datetime.now().strftime("%d.%m.%Y")
            else:
                combined_value += merge_data.get(tag, tag)
        merge_data[combo['name']] = combined_value

    return merge_data

def guess_type(value: str):
    """Try to guess if the string should be int, float, date, or text."""

    val = str(value).strip()

    # Try integer (only digits)
    if val.isdigit():
        try:
            return int(val)
        except ValueError:
            pass

    # Try float (handle decimal commas and points)
    try:
        if "." in val or "," in val:
            # Replace comma with dot for float conversion
            val_float = val.replace(",", ".")
            return float(val_float)
    except ValueError:
        pass

    # Try date with common formats
    date_formats = [
        "%Y-%m-%d",   # 2023-08-10
        "%d.%m.%Y",   # 10.08.2023
        "%d/%m/%Y",   # 10/08/2023
        "%m/%d/%Y",   # 08/10/2023
        "%d-%m-%Y",   # 10-08-2023
        "%Y/%m/%d",   # 2023/08/10
        # add more if needed
    ]

    for fmt in date_formats:
        try:
            dt = datetime.strptime(val, fmt)
            return dt.date()  # or dt if you want datetime
        except ValueError:
            continue

    # If all else fails, return original string
    return val

def submit_and_save():
    """Main function to generate documents after validation."""
    global dynamic_frame

    # Check if any fields have been created
    if not dynamic_frame or not dynamic_frame.winfo_children():
        messagebox.showinfo(
            "Информация",
            "Пожалуйста, создайте хотя бы одно поле в Конструкторе перед формированием документов."
        )
        return

    try:
        # Determine the base directory for both dev and PyInstaller environments
        if getattr(sys, 'frozen', False):
            BASE_DIR = sys._MEIPASS
        else:
            BASE_DIR = os.path.dirname(os.path.abspath(__file__))

        # Define source and output directories
        source_dir = os.path.join(BASE_DIR, 'documents', 'template')
        output_dir = os.path.join(BASE_DIR, 'documents', 'processed')

        # Ensure output directory exists
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # Get merge data
        common_data = get_common_merge_data()

        # Apply rules
        rules = load_json(RULES_CONFIG_PATH, 'rules_config')
        if not isinstance(rules, list):
            rules = []

        for rule in rules:
            conditions = rule.get('conditions', [])
            behaviors = rule.get('behaviors', [])

            if not conditions:
                common_data = apply_behaviors(behaviors, common_data)
                continue

            all_conditions_met = all(evaluate_condition(cond, common_data) for cond in conditions)

            if all_conditions_met:
                non_cleaner_behaviors = [
                    b for b in behaviors if b.get('condition') != 'очистить при не выполнении'
                ]
                if non_cleaner_behaviors:
                    common_data = apply_behaviors(non_cleaner_behaviors, common_data)
            else:
                cleaner_behavior = next(
                    (b for b in behaviors if b.get('condition') == 'очистить при не выполнении'), None
                )
                if cleaner_behavior:
                    common_data = apply_behaviors([cleaner_behavior], common_data)

        # Timestamp for output filenames
        timestamp = datetime.now().strftime("%m%d%H%M%S")

        # Process DOCX files
        docx_files = [f for f in os.listdir(source_dir) if f.endswith('.docx')]
        for docx_file in docx_files:
            try:
                file_name_without_ext, _ = os.path.splitext(docx_file)
                new_file_name = f"{file_name_without_ext}_{timestamp}.docx"
                output_path = os.path.join(output_dir, new_file_name)

                if os.path.exists(output_path):
                    if not messagebox.askyesno(
                        "Перезапись файла",
                        f"Файл '{new_file_name}' уже существует. Хотите перезаписать?"
                    ):
                        continue

                document = MailMerge(os.path.join(source_dir, docx_file))
                merge_fields_in_doc = document.get_merge_fields()
                filtered_data = {key: common_data.get(key, '') for key in merge_fields_in_doc}

                document.merge(**filtered_data)
                document.write(output_path)
                document.close()
            except Exception as e:
                print(f"Error processing {docx_file}: {e}")

        # Process XLS/XLSX files
        xls_files = [f for f in os.listdir(source_dir) if f.endswith(('.xls', '.xlsx'))]
        for xls_file in xls_files:
            try:
                file_name_without_ext, _ = os.path.splitext(xls_file)
                new_file_name = f"{file_name_without_ext}_{timestamp}.xlsx"
                output_path = os.path.join(output_dir, new_file_name)

                if os.path.exists(output_path):
                    if not messagebox.askyesno(
                        "Перезапись файла",
                        f"Файл '{new_file_name}' уже существует. Хотите перезаписать?"
                    ):
                        continue

                wb = openpyxl.load_workbook(os.path.join(source_dir, xls_file))
                sheet = wb.active

                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and str(cell.value) in common_data:
                            new_val = guess_type(common_data[str(cell.value)])

                            # Assign type & number format
                            if isinstance(new_val, (int, float)):
                                cell.value = new_val
                                cell.number_format = numbers.FORMAT_GENERAL
                            elif isinstance(new_val, date):
                                cell.value = new_val
                                cell.number_format = 'DD.MM.YYYY'
                            else:
                                cell.value = str(new_val)
                                cell.number_format = numbers.FORMAT_GENERAL

                wb.save(output_path)

            except Exception as e:
                print(f"Error processing {xls_file}: {e}")

        messagebox.showinfo(title="Успех!", message="Документы были успешно сформированы.")

    except Exception as e:
        messagebox.showerror(
            title="Ошибка!",
            message=f"Произошла ошибка при формировании документов: {e}"
        )


def import_fields():
    global dynamic_frame

    # Prompt user for confirmation
    if not messagebox.askyesno("Подтверждение", "Вы уверены, что хотите импортировать поля из файла 'field_import'?"):
        return

    # Determine the base directory
    if getattr(sys, 'frozen', False):
        # PyInstaller
        BASE_DIR = sys._MEIPASS
    else:
        # Regular Python script
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))

    # Define the folder for your import files
    IMPORT_FLD = os.path.join(BASE_DIR, 'import_fld')

    # Define file paths relative to the import folder
    xlsx_path = os.path.join(IMPORT_FLD, "field_import.xlsx")
    xls_path = os.path.join(IMPORT_FLD, "field_import.xls")

    # Check for file existence, prioritizing .xlsx
    if os.path.exists(xlsx_path):
        file_path = xlsx_path
    elif os.path.exists(xls_path):
        file_path = xls_path
    else:
        messagebox.showerror("Ошибка", "Ошибка: Файл 'field_import' не найден.")
        return

    try:
        # Load Excel file
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Validate at least two columns
        if sheet.max_column < 2:
            messagebox.showerror("Ошибка",
                                 "Ошибка: Файл 'field_import' поврежден или не содержит двух столбцов с данными.")
            return

        # Initialize counters and tracking
        fields_created = 0
        values_imported = 0
        processed_names = set()  # Track field names (case-insensitive)
        value_map = {}  # Store values for reapplying after refresh

        # Load all config files
        fields_config = load_json(FIELDS_CONFIG_PATH, [])
        combobox_regular = load_json(COMBOBOX_REGULAR_PATH, [])
        combobox_mainkey = load_json(COMBOBOX_MAINKEY_PATH, [])
        combination_config = load_json(COMBINATION_CONFIG_PATH, [])

        # Process Excel rows
        for row in sheet.iter_rows(values_only=True):
            field_name = row[0]
            field_value = row[1]

            # Skip rows with empty field name
            if not field_name or not str(field_name).strip():
                continue

            # Check for duplicate field name (case-insensitive)
            field_name_lower = str(field_name).lower()
            if field_name_lower in processed_names:
                continue
            processed_names.add(field_name_lower)

            # Store value for later use (even if empty)
            value_map[field_name_lower] = str(field_value) if field_value is not None else ""

            # Check if field exists in any config (case-insensitive)
            exists = False
            tag_type = None
            for config in [fields_config, combobox_regular, combobox_mainkey, combination_config]:
                for item in config:
                    if item.get("name", "").lower() == field_name_lower:
                        exists = True
                        tag_type = item.get("tag_type")
                        break
                if exists:
                    break

            if exists:
                # If field exists and is type "поле", update its value
                if tag_type == "поле":
                    found = False
                    for widget in dynamic_frame.winfo_children():
                        if hasattr(widget, "_name") and widget._name.lower() == field_name_lower and isinstance(widget,
                                                                                                                tk.Entry):
                            widget.delete(0, tk.END)
                            widget.insert(0, value_map[field_name_lower])
                            values_imported += 1
                            found = True
                            break
                    if not found:
                        print(f"Warning: No Entry widget found for existing field '{field_name}'")
                # Skip non-"поле" fields
                continue

            # Create new field if unique
            if field_name_lower not in [item["name"].lower() for config in
                                        [fields_config, combobox_regular, combobox_mainkey, combination_config] for item
                                        in config]:
                fields_config.append({
                    "name": str(field_name),
                    "type": "текст",
                    "tag_type": "поле"
                })
                # Add widget to dynamic_frame with value
                add_dynamic_widget(str(field_name), "текст", "поле", value_map[field_name_lower])
                fields_created += 1
                values_imported += 1  # Count value even if empty, as it's applied

        # Save updated fields config
        save_json(FIELDS_CONFIG_PATH, fields_config)

        # Refresh UI and reapply values
        refresh_main_and_constructor()

        # Reapply values to all Entry widgets
        for widget in dynamic_frame.winfo_children():
            if hasattr(widget, "_name") and isinstance(widget, tk.Entry):
                field_name_lower = widget._name.lower()
                if field_name_lower in value_map:
                    widget.delete(0, tk.END)
                    widget.insert(0, value_map[field_name_lower])

        # Show feedback
        messagebox.showinfo("Импорт завершен",
                            f"{fields_created} полей было импортировано, {values_imported} значений было импортировано.")

    except Exception as e:
        messagebox.showerror("Ошибка", "Ошибка: Файл 'field_import' поврежден или не содержит двух столбцов с данными.")
        print(f"Import error: {str(e)}")


def get_next_grid_position():
    """Calculates the next grid position (row, column) based on a 30-widget-per-column rule."""
    global dynamic_frame
    if not dynamic_frame:
        return 0, 0

    # Each widget consists of a Label and an Entry/Combobox, so we count pairs.
    widget_count = len(dynamic_frame.winfo_children()) // 2

    row = widget_count % 30
    col_group = widget_count // 30

    # Each column group for widgets takes 2 grid columns (one for label, one for widget)
    base_col = col_group * 2

    return row, base_col


def add_dynamic_widget(name, data_type, tag_type, values=None, main_key_data=None, initial_value=None):
    """Adds a new widget to the dynamic_frame, optionally with an initial value."""
    global dynamic_frame, checkbox_vars, main_key_selections

    row, base_col = get_next_grid_position()

    label = tk.Label(dynamic_frame, text=f"{name}:")
    label._name = f"{name}l"
    label.grid(row=row, column=base_col, padx=5, pady=2, sticky="e")

    if tag_type == "поле":
        entry = tk.Entry(dynamic_frame, width=25)
        entry._name = name
        if initial_value is not None:
            entry.insert(0, str(initial_value))
        entry.grid(row=row, column=base_col + 1, padx=5, pady=2, sticky="w")

    elif tag_type == "чекбокс":
        var = tk.IntVar()
        if initial_value is not None:
            var.set(initial_value)
        checkbox_vars[name] = var
        checkbox = ttk.Checkbutton(dynamic_frame, variable=var)
        checkbox._name = name
        checkbox.grid(row=row, column=base_col + 1, padx=5, pady=2, sticky="w")

    elif tag_type == "комбобокс":  # Regular combobox
        combobox = tkentrycomplete.Combobox(dynamic_frame, values=values, width=22)
        combobox._name = name
        if initial_value is not None:
            combobox.set(initial_value)
        combobox.set_completion_list({v: {} for v in values})
        combobox.grid(row=row, column=base_col + 1, padx=5, pady=2, sticky="w")

    elif tag_type == "список":  # Main-key combobox
        var = tk.StringVar()
        if initial_value is not None:
            var.set(initial_value)
        combobox = tkentrycomplete.Combobox(dynamic_frame, values=values, textvariable=var, width=22)
        combobox._name = name
        combobox.set_completion_list(main_key_data)

        def on_select(event, widget_name=name, data=main_key_data):
            selected_key = event.widget.get()
            if selected_key in data:
                main_key_selections[widget_name] = data[selected_key]
            elif widget_name in main_key_selections:
                del main_key_selections[widget_name]

        combobox.bind('<<ComboboxSelected>>', on_select)
        combobox.bind('<FocusOut>', on_select)
        combobox.bind('<Return>', on_select)
        combobox.grid(row=row, column=base_col + 1, padx=5, pady=2, sticky="w")


def load_all_dynamic_widgets(initial_state=None):
    """Loads all configured UI elements, optionally applying an initial state."""
    if initial_state is None:
        initial_state = {}

    # --- (The rest of the function for loading and sorting configs remains the same) ---
    fields_list = []
    comboboxes_list = []
    checkboxes_list = []

    # --- Load Fields & Checkboxes ---
    fields_config = load_json(FIELDS_CONFIG_PATH, 'fields_config')
    for field in fields_config:
        tag_type = field.get('tag_type', 'поле')
        if tag_type == 'чекбокс':
            checkboxes_list.append((field['name'], field['type'], tag_type, None, None))
        else:  # Regular text field
            fields_list.append((field['name'], field['type'], tag_type, None, None))

    # --- Load Regular Comboboxes ---
    regular_combos = load_json(COMBOBOX_REGULAR_PATH, 'combobox_regular')
    for combo in regular_combos:
        comboboxes_list.append((combo['name'], combo['type'], 'комбобокс', combo['values'], None))

    # --- Load Main-Key Comboboxes ---
    mainkey_combos = load_json(COMBOBOX_MAINKEY_PATH, 'combobox_mainkey')
    for combo in mainkey_combos:
        values = [list(mk.keys())[0] for mk in combo['main_keys']]
        data_dict = {list(mk.keys())[0]: list(mk.values())[0] for mk in combo['main_keys']}
        comboboxes_list.append((combo['name'], combo['type'], 'список', values, data_dict))

    # --- Sort each group alphabetically by name ---
    fields_list.sort(key=lambda x: x[0].lower())
    comboboxes_list.sort(key=lambda x: x[0].lower())
    checkboxes_list.sort(key=lambda x: x[0].lower())

    # --- Combine groups in desired order ---
    ordered_widgets = fields_list + comboboxes_list + checkboxes_list

    # --- Place widgets in order, passing the saved value ---
    for idx, (name, data_type, tag_type, values, main_key_data) in enumerate(ordered_widgets):
        # ... (Separator logic remains the same) ...

        # Get the saved value for this widget, if it exists
        saved_value = initial_state.get(name)

        # Pass the saved value to the creation function
        add_dynamic_widget(name, data_type, tag_type, values, main_key_data, initial_value=saved_value)



def get_all_tags_for_constructor():
    """Gathers a flat list of all tags for use in constructor UI elements."""
    all_tags = []

    # Add fields, checkboxes, and regular comboboxes
    all_tags.extend([f['name'] for f in load_json(FIELDS_CONFIG_PATH, 'fields_config')])
    all_tags.extend([c['name'] for c in load_json(COMBOBOX_REGULAR_PATH, 'combobox_regular')])

    # Add main keys and their subkeys
    mainkey_combos = load_json(COMBOBOX_MAINKEY_PATH, 'combobox_mainkey')
    for combo in mainkey_combos:
        # Add the main key itself
        all_tags.append(combo['name'])
        # Add each subkey
        for mk_dict in combo.get('main_keys', []):
            # Assumes each main_keys item is a dictionary with one key (main_key)
            subkeys_dict = list(mk_dict.values())[0]
            for subkey_name in subkeys_dict.keys():
                all_tags.append(subkey_name)

    # Add combination tags
    combination_config = load_json(COMBINATION_CONFIG_PATH, 'combination_config')
    all_tags.extend([c['name'] for c in combination_config])

    return sorted(list(set(all_tags)))  # Return sorted unique tags

# --- Constructor Window and its Helpers ---

def sort_column(treeview, col, reverse):
    """Sorts a Treeview column."""
    items = [(treeview.set(item, col), item) for item in treeview.get_children('')]
    items.sort(key=lambda x: str(x[0]).lower(), reverse=reverse)
    for index, (_, item) in enumerate(items):
        treeview.move(item, '', index)
    treeview.heading(col, command=lambda: sort_column(treeview, col, not reverse))


def update_rules_listbox(rules, listbox):
    # Calculate maximum number of conditions and behaviors, with minimum 1 each
    max_conditions = max((len(rule.get('conditions', [])) for rule in rules), default=1) if rules else 1
    max_behaviors = max((len(rule.get('behaviors', [])) for rule in rules), default=1) if rules else 1

    # Define columns: always include "Имя", "Условие 1", "Поведение 1"
    required_columns = ["Имя"] + [f"Условие {i + 1}" for i in range(max_conditions)] + [f"Поведение {i + 1}" for i in
                                                                                           range(max_behaviors)]

    # Always set columns to ensure default columns are present
    listbox["columns"] = required_columns
    for col in required_columns:
        listbox.heading(col, text=col, anchor="center")
        listbox.column(col, anchor="center", stretch=True)

    # Clear existing items
    for item in listbox.get_children():
        listbox.delete(item)

    # Populate listbox
    for rule in rules:
        name = rule.get('name', '')
        conditions = [f"{c['tag']} {c['condition']} {c['rule']}" for c in rule.get('conditions', [])]
        behaviors = [f"{b['tag']} {b['condition']} {b['rule']}" for b in rule.get('behaviors', [])]
        # Align values with columns: name, conditions, padding, behaviors, padding
        values = [name] + conditions + [""] * (max_conditions - len(conditions)) + behaviors + [""] * (
            max_behaviors - len(behaviors))
        listbox.insert("", "end", values=values)


def open_constructor_window():
    """Opens the main constructor window for managing tags and rules."""
    constructor_window = tk.Toplevel(window)
    constructor_window.title("Конструктор")
    constructor_window.geometry("1200x600")
    constructor_window.focus_set()

    # This intercepts the "X" button press and calls our custom function
    constructor_window.protocol("WM_DELETE_WINDOW", lambda: on_constructor_close(constructor_window))

    notebook = ttk.Notebook(constructor_window)
    notebook.pack(pady=10, padx=10, fill="both", expand=True)

    # --- Tags Tab ---
    tags_tab = ttk.Frame(notebook)
    notebook.add(tags_tab, text='Теги')

    tags_list_frame = tk.Frame(tags_tab)
    tags_list_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
    tags_buttons_frame = tk.Frame(tags_tab)
    tags_buttons_frame.grid(row=0, column=1, padx=5, pady=5, sticky="ns")
    tags_tab.grid_rowconfigure(0, weight=1)
    tags_tab.grid_columnconfigure(0, weight=1)

    tags_listbox = ttk.Treeview(tags_list_frame, columns=("Имя", "Тип ввода", "Тип тега"), show="headings", height=20)
    tags_listbox.heading("Имя", text="Имя", command=lambda: sort_column(tags_listbox, "Имя", False))
    tags_listbox.heading("Тип ввода", text="Тип ввода", command=lambda: sort_column(tags_listbox, "Тип ввода", False))
    tags_listbox.heading("Тип тега", text="Тип тега", command=lambda: sort_column(tags_listbox, "Тип тега", False))
    tags_listbox.column("Имя", width=250, anchor="center")
    tags_listbox.column("Тип ввода", width=120, anchor="center")
    tags_listbox.column("Тип тега", width=120, anchor="center")

    scrollbar_tags = ttk.Scrollbar(tags_list_frame, orient="vertical", command=tags_listbox.yview)
    tags_listbox.configure(yscrollcommand=scrollbar_tags.set)
    tags_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar_tags.pack(side=tk.RIGHT, fill=tk.Y)

    tk.Button(tags_buttons_frame, text="Новый", width=14, command=lambda: open_new_tag_window(tags_listbox, constructor_window)).pack(
        side=TOP, pady=2)
    tk.Button(tags_buttons_frame, text="Редактировать", width=14,
              command=lambda: open_edit_tag_window(tags_listbox, constructor_window)).pack(side=TOP, pady=2)
    tk.Button(tags_buttons_frame, text="Удалить", width=14, command=lambda: delete_tag(tags_listbox, constructor_window)).pack(side=TOP,
                                                                                                           pady=2)
    populate_tags_listbox(tags_listbox)

    # --- Rules Tab ---
    rules_tab = ttk.Frame(notebook)
    notebook.add(rules_tab, text='Правила')

    rules_list_frame = tk.Frame(rules_tab)
    rules_list_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
    rules_buttons_frame = tk.Frame(rules_tab)
    rules_buttons_frame.grid(row=0, column=1, padx=5, pady=5, sticky="ns")
    rules_tab.grid_rowconfigure(0, weight=1)
    rules_tab.grid_columnconfigure(0, weight=1)

    rules_listbox = ttk.Treeview(rules_list_frame, show="headings", height=20)

    v_scrollbar_rules = ttk.Scrollbar(rules_list_frame, orient="vertical", command=rules_listbox.yview)
    h_scrollbar_rules = ttk.Scrollbar(rules_list_frame, orient="horizontal", command=rules_listbox.xview)
    rules_listbox.configure(yscrollcommand=v_scrollbar_rules.set, xscrollcommand=h_scrollbar_rules.set)

    rules_listbox.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
    v_scrollbar_rules.pack(side=tk.RIGHT, fill=tk.Y, before=rules_listbox)
    h_scrollbar_rules.pack(side=tk.BOTTOM, fill=tk.X)

    tk.Button(rules_buttons_frame, text="Создать", width=14,
              command=lambda: open_create_rule_window(rules_listbox, constructor_window)).pack(side=TOP, pady=2)
    tk.Button(rules_buttons_frame, text="Изменить", width=14,
              command=lambda: open_edit_rule_window(rules_listbox, constructor_window)).pack(side=TOP, pady=2)
    tk.Button(rules_buttons_frame, text="Удалить", width=14, command=lambda: delete_rule(rules_listbox, constructor_window)).pack(side=TOP,
                                                                                                              pady=2)

    rules = load_json(RULES_CONFIG_PATH, 'rules_config')
    update_rules_listbox(rules, rules_listbox)


# All other constructor helper functions (open_new_tag_window, open_field_window, etc.) need to be included here.
# For brevity, I will add the main ones back. The complex ones like rules will need to be added from original.
# The following implementations are from the original file, adapted for the new structure.

def open_new_tag_window(listbox, parent_window):
    """Window to choose what kind of new tag to create."""
    new_window = tk.Toplevel(parent_window)
    new_window.title("Новый тег")
    new_window.geometry("200x250")
    new_window.resizable(False, False)
    new_window.focus_set()
    new_window.grab_set()

    btn_frame = tk.Frame(new_window)
    btn_frame.pack(pady=10, expand=True)

    # The listbox needs to be passed to refresh it upon creation
    tk.Button(btn_frame, text="ПОЛЕ", width=15, height=2,
              command=lambda: [new_window.destroy(), open_field_window(listbox, None, parent_window)]).pack(pady=5)
    tk.Button(btn_frame, text="СПИСОК", width=15, height=2,
              command=lambda: [new_window.destroy(), open_list_window(listbox, None, parent_window)]).pack(pady=5)
    tk.Button(btn_frame, text="ЧЕКБОКС", width=15, height=2,
              command=lambda: [new_window.destroy(), open_checkbox_window(listbox, None, parent_window)]).pack(pady=5)
    tk.Button(btn_frame, text="СОЧЕТАНИЕ", width=15, height=2, command=lambda: [new_window.destroy(), open_combination_window(listbox, None, parent_window)]).pack(pady=5)


def open_field_window(listbox, item_to_edit, parent_window):
    """Window to create or edit a simple 'Field' (Entry widget)."""
    is_edit = item_to_edit is not None
    title = "Редактировать поле" if is_edit else "Создание поля"

    field_window = tk.Toplevel(parent_window)
    field_window.title(title)
    field_window.geometry("300x180")
    field_window.resizable(False, False)
    field_window.focus_set()
    field_window.grab_set()

    old_name = ""
    if is_edit:
        old_name, old_type, _ = listbox.item(item_to_edit)['values']
        name_var = tk.StringVar(value=old_name)
        type_var = tk.StringVar(value=old_type)
    else:
        name_var = tk.StringVar()
        type_var = tk.StringVar(value="текст")

    tk.Label(field_window, text="Имя поля:").pack(pady=(10, 0))
    name_entry = tk.Entry(field_window, textvariable=name_var, width=30)
    name_entry.pack(pady=5, padx=10)
    name_entry.focus()

    tk.Label(field_window, text="Тип данных:").pack(pady=(10, 0))
    type_frame = tk.Frame(field_window)
    type_frame.pack(pady=5)
    types = [("текст", "текст"), ("числа", "числа"), ("дата", "дата")]
    for text, value in types:
        tk.Radiobutton(type_frame, text=text, value=value, variable=type_var).pack(side=LEFT, padx=5)

    def save_field():
        name = name_var.get().strip()
        data_type = type_var.get()
        if not name:
            messagebox.showwarning("Ошибка", "Введите имя поля.", parent=field_window)
            return

        all_configs = (load_json(path, '') for path in
                       [FIELDS_CONFIG_PATH, COMBOBOX_REGULAR_PATH, COMBOBOX_MAINKEY_PATH, COMBINATION_CONFIG_PATH])
        all_names = {item['name'] for config in all_configs for item in config}

        if name != old_name and name in all_names:
            messagebox.showwarning("Ошибка", f"Имя '{name}' уже используется.", parent=field_window)
            return

        fields_config = load_json(FIELDS_CONFIG_PATH, 'fields_config')
        if is_edit:
            for field in fields_config:
                if field['name'] == old_name:
                    field['name'] = name
                    field['type'] = data_type
                    break
        else:
            fields_config.append({'name': name, 'type': data_type, 'tag_type': 'поле'})

        save_json(FIELDS_CONFIG_PATH, fields_config)
        refresh_all_windows(listbox)
        field_window.destroy()

    btn_frame = tk.Frame(field_window)
    btn_frame.pack(pady=10)
    tk.Button(btn_frame, text="ОК", width=10, command=save_field).pack(side=LEFT, padx=5)
    tk.Button(btn_frame, text="ОТМЕНА", width=10, command=field_window.destroy).pack(side=LEFT, padx=5)


def open_checkbox_window(listbox, item_to_edit, parent_window):
    """Window to create or edit a 'Checkbox'."""
    is_edit = item_to_edit is not None
    title = "Редактировать чекбокс" if is_edit else "Создание чекбокса"

    cb_window = tk.Toplevel(parent_window)
    cb_window.title(title)
    cb_window.geometry("300x120")
    cb_window.resizable(False, False)
    cb_window.focus_set()
    cb_window.grab_set()

    old_name = ""
    if is_edit:
        old_name, _, _ = listbox.item(item_to_edit)['values']
        name_var = tk.StringVar(value=old_name)
    else:
        name_var = tk.StringVar()

    tk.Label(cb_window, text="Имя чекбокса:").pack(pady=(10, 0))
    name_entry = tk.Entry(cb_window, textvariable=name_var, width=30)
    name_entry.pack(pady=5, padx=10)
    name_entry.focus()

    def save_checkbox():
        name = name_var.get().strip()
        if not name:
            messagebox.showwarning("Ошибка", "Введите имя чекбокса.", parent=cb_window)
            return

        all_configs = (load_json(path, '') for path in
                       [FIELDS_CONFIG_PATH, COMBOBOX_REGULAR_PATH, COMBOBOX_MAINKEY_PATH, COMBINATION_CONFIG_PATH])
        all_names = {item['name'] for config in all_configs for item in config}

        if name != old_name and name in all_names:
            messagebox.showwarning("Ошибка", f"Имя '{name}' уже используется.", parent=cb_window)
            return

        fields_config = load_json(FIELDS_CONFIG_PATH, 'fields_config')
        if is_edit:
            for field in fields_config:
                if field['name'] == old_name:
                    field['name'] = name
                    break
        else:
            fields_config.append({'name': name, 'type': 'чекбокс', 'tag_type': 'чекбокс'})

        save_json(FIELDS_CONFIG_PATH, fields_config)
        refresh_all_windows(listbox)
        cb_window.destroy()

    btn_frame = tk.Frame(cb_window)
    btn_frame.pack(pady=10)
    tk.Button(btn_frame, text="ОК", width=10, command=save_checkbox).pack(side=LEFT, padx=5)
    tk.Button(btn_frame, text="ОТМЕНА", width=10, command=cb_window.destroy).pack(side=LEFT, padx=5)


# --- REPLACE THE OLD open_list_window FUNCTION WITH THIS ---

def open_list_window(listbox, item_to_edit, parent_window):
    """
    Full implementation for CREATING and EDITING lists.
    Now for 'список' type in edit mode, uses a main key selector instead of showing all main keys at once.
    """
    is_edit = item_to_edit is not None
    title = "Редактировать список" if is_edit else "Создание списка"

    list_window = tk.Toplevel(parent_window)
    list_window.title(title)
    list_window.resizable(False, False)
    list_window.focus_set()
    list_window.grab_set()

    # --- Initial Variable Setup ---
    name_var = tk.StringVar()
    main_key_var = tk.IntVar()
    sets_list = []
    old_name = ""
    original_tag_type = ""

    # --- Data Loading for Edit Mode ---
    if is_edit:
        old_name, _, tag_type = listbox.item(item_to_edit)['values']
        original_tag_type = tag_type
        name_var.set(old_name)

        if tag_type == 'комбобокс':
            main_key_var.set(0)
            config = load_json(COMBOBOX_REGULAR_PATH, '')
            list_data = next((item for item in config if item['name'] == old_name), None)
            if list_data:
                sets_list.append({'initial_values': list_data.get('values', [])})

        elif tag_type == 'список':
            main_key_var.set(1)
            config = load_json(COMBOBOX_MAINKEY_PATH, '')
            list_data = next((item for item in config if item['name'] == old_name), None)
            if list_data:
                for main_key_dict in list_data.get('main_keys', []):
                    for main_key, sub_dict in main_key_dict.items():
                        new_set = {
                            "main_key": tk.StringVar(value=main_key),
                            "key_values": [{"key": tk.StringVar(value=k), "value": tk.StringVar(value=v)}
                                           for k, v in sub_dict.items()]
                        }
                        sets_list.append(new_set)
    else:
        sets_list.append({"main_key": tk.StringVar(), "key_values": [{"key": tk.StringVar(), "value": tk.StringVar()}]})

    # --- Helper Functions ---
    def add_key_value_row():
        for s in sets_list:
            s["key_values"].append({"key": tk.StringVar(), "value": tk.StringVar()})
        refresh_table()

    def add_key_value_to_specific_set(set_index):
        if 0 <= set_index < len(sets_list):
            sets_list[set_index]["key_values"].append({"key": tk.StringVar(), "value": tk.StringVar()})
            refresh_table()

    def add_set():
        # Get subkey names from the first set (ignores values, just keeps keys)
        if sets_list:
            key_structure = [kv['key'].get() for kv in sets_list[0]['key_values']]
        else:
            key_structure = []

        # Create new key-value dicts: same keys, empty values
        new_key_values = [{"key": tk.StringVar(value=k), "value": tk.StringVar()} for k in key_structure]

        sets_list.append({
            "main_key": tk.StringVar(),
            "key_values": new_key_values
        })

        refresh_table()

    def refresh_table():
        for widget in table_frame.winfo_children():
            widget.destroy()

        if not main_key_var.get():
            # --- Simple list mode ---
            simple_frame = tk.Frame(table_frame)
            simple_frame.pack(fill="both", expand=True)
            tk.Label(simple_frame, text="Значения (каждое с новой строки):").pack(anchor="w")
            text_area = Text(simple_frame, width=60, height=15)
            text_area.pack(fill="both", expand=True, padx=5, pady=5)
            if is_edit and sets_list and 'initial_values' in sets_list[0]:
                text_area.insert('1.0', '\n'.join(sets_list[0]['initial_values']))
            if not is_edit:
                tk.Button(simple_frame, text="Импорт из Excel", command=import_from_excel).pack(pady=5)
            sets_list[0]['widget_ref'] = text_area



        else:

            if is_edit:

                # --- Main key edit mode with selector ---

                selector_frame = tk.Frame(table_frame)

                selector_frame.pack(fill="x", padx=10, pady=5)

                tk.Label(selector_frame, text="Выберите главный ключ:", anchor="w").pack(side="left")

                main_keys_list = [s["main_key"].get() for s in sets_list if s["main_key"].get().strip()]

                selected_main_key = tk.StringVar()

                main_key_combo = ttk.Combobox(selector_frame, values=main_keys_list,

                                              textvariable=selected_main_key, state="readonly", width=25)

                main_key_combo.pack(side="left", padx=5)

                subkeys_frame = tk.Frame(table_frame)

                subkeys_frame.pack(fill="x", padx=10, pady=5)

                def show_subkeys_for_selected(*args):

                    for w in subkeys_frame.winfo_children():
                        w.destroy()

                    key = selected_main_key.get()

                    set_index = next((i for i, s in enumerate(sets_list) if s["main_key"].get() == key), None)

                    if set_index is None:
                        return

                    s = sets_list[set_index]

                    tk.Label(subkeys_frame, text=f"Главный ключ:", anchor="w").grid(row=0, column=0, sticky="w", pady=2)

                    tk.Entry(subkeys_frame, textvariable=s["main_key"], width=20).grid(row=0, column=1, sticky="w",
                                                                                       padx=5)

                    for i, kv in enumerate(s["key_values"], start=1):
                        tk.Label(subkeys_frame, text=f"Ключ {i}:", anchor="w").grid(row=i, column=0, sticky="w", pady=2)

                        tk.Entry(subkeys_frame, textvariable=kv["key"], width=20).grid(row=i, column=1, sticky="w",
                                                                                       padx=5)

                        tk.Label(subkeys_frame, text="Значение:", anchor="w").grid(row=i, column=2, sticky="w", padx=5)

                        tk.Entry(subkeys_frame, textvariable=kv["value"], width=20).grid(row=i, column=3, sticky="w",
                                                                                         padx=5)

                    tk.Button(subkeys_frame, text="Добавить строку",

                              command=lambda: add_key_value_to_specific_set(set_index)).grid(

                        row=len(s["key_values"]) + 1, column=0, pady=5)

                main_key_combo.bind("<<ComboboxSelected>>", show_subkeys_for_selected)

                if main_keys_list:
                    selected_main_key.set(main_keys_list[0])

                    show_subkeys_for_selected()

                control_frame = tk.Frame(table_frame)

                control_frame.pack(pady=10)

                tk.Button(control_frame, text="Добавить главный ключ", command=add_set).pack(side=LEFT, padx=5)


            else:

                # --- Main key creation mode ---

                for set_index, s in enumerate(sets_list):

                    set_frame = ttk.LabelFrame(table_frame, text=f"Набор {set_index + 1}")

                    set_frame.pack(fill="x", expand=True, padx=5, pady=5)

                    keys_frame = tk.Frame(set_frame)

                    keys_frame.pack(fill="x", padx=10, pady=2)

                    tk.Label(keys_frame, text="Главный ключ:", anchor="w").grid(row=0, column=0, sticky="w", pady=2)

                    tk.Entry(keys_frame, textvariable=s["main_key"], width=25).grid(row=0, column=1, sticky="w", padx=5)

                    for i, kv in enumerate(s["key_values"]):
                        row_num = i + 1

                        tk.Label(keys_frame, text=f"Ключ {row_num}:", anchor="w").grid(row=row_num, column=0,
                                                                                       sticky="w", pady=2)

                        tk.Entry(keys_frame, textvariable=kv["key"], width=20).grid(row=row_num, column=1, sticky="w",
                                                                                    padx=5)

                        tk.Label(keys_frame, text="Значение:", anchor="w").grid(row=row_num, column=2, sticky="w",
                                                                                padx=5)

                        tk.Entry(keys_frame, textvariable=kv["value"], width=20).grid(row=row_num, column=3, sticky="w",
                                                                                      padx=5)

                    tk.Button(set_frame, text="Добавить строку в этот набор",

                              command=lambda si=set_index: add_key_value_to_specific_set(si)).pack(pady=5)

                control_frame = tk.Frame(table_frame)

                control_frame.pack(pady=10)

                tk.Button(control_frame, text="Строка", command=add_key_value_row).pack(side=LEFT, padx=5)

                tk.Button(control_frame, text="Добавить", command=add_set).pack(side=LEFT, padx=5)

                if not is_edit:
                    tk.Button(control_frame, text="Импорт", command=import_from_excel).pack(side=LEFT, padx=5)

    def save_combobox():
        name = name_var.get().strip()
        if not name:
            messagebox.showwarning("Ошибка", "Введите имя списка.", parent=list_window)
            return
        all_configs = (load_json(path, '') for path in
                       [FIELDS_CONFIG_PATH, COMBOBOX_REGULAR_PATH, COMBOBOX_MAINKEY_PATH, COMBINATION_CONFIG_PATH])
        all_names = {item['name'] for config in all_configs for item in config}
        if name != old_name and name in all_names:
            messagebox.showwarning("Ошибка", f"Имя '{name}' уже используется.", parent=list_window)
            return

        if not main_key_var.get():
            widget = sets_list[0].get('widget_ref')
            if not widget: return
            values = [v.strip() for v in widget.get("1.0", "end-1c").split('\n') if v.strip()]
            combo_data = {"name": name, "type": "текст", "tag_type": "комбобокс", "values": sorted(values)}
            config_path, other_config_path = COMBOBOX_REGULAR_PATH, COMBOBOX_MAINKEY_PATH
        else:
            main_keys = []
            for s in sets_list:
                mk = s["main_key"].get().strip()
                if not mk: continue
                kv_pairs = {kv["key"].get().strip(): kv["value"].get().strip()
                            for kv in s["key_values"] if kv["key"].get().strip()}
                if kv_pairs:
                    main_keys.append({mk: kv_pairs})
            combo_data = {"name": name, "type": "текст", "tag_type": "список", "main_keys": main_keys}
            config_path, other_config_path = COMBOBOX_MAINKEY_PATH, COMBOBOX_REGULAR_PATH

        if is_edit:
            original_file_path = COMBOBOX_REGULAR_PATH if original_tag_type == 'комбобокс' else COMBOBOX_MAINKEY_PATH
            cfg = load_json(original_file_path, '')
            cfg = [item for item in cfg if item.get('name') != old_name]
            save_json(original_file_path, cfg)
            if config_path != original_file_path:
                other_cfg = load_json(other_config_path, '')
                other_cfg = [item for item in other_cfg if item.get('name') != old_name]
                save_json(other_config_path, other_cfg)

        final_config = load_json(config_path, '')
        final_config.append(combo_data)
        save_json(config_path, final_config)
        refresh_all_windows(listbox)
        list_window.destroy()

    def import_from_excel():
        # Determine the base directory
        if getattr(sys, 'frozen', False):
            # PyInstaller
            BASE_DIR = sys._MEIPASS
        else:
            # Regular Python script
            BASE_DIR = os.path.dirname(os.path.abspath(__file__))

        # Define the folder for your import files
        IMPORT_FLD = os.path.join(BASE_DIR, 'import_fld')

        # Define the file path relative to the import folder
        import_path = os.path.join(IMPORT_FLD, "import.xlsx")
        if not os.path.exists(import_path):
            messagebox.showwarning("Ошибка", f"Файл {import_path} не найден", parent=list_window)
            return
        if not name_var.get().strip():
            messagebox.showwarning("Ошибка", "Введите имя списка перед импортом", parent=list_window)
            return
        try:
            wb = openpyxl.load_workbook(import_path)
            sheet = wb.active
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл {import_path}. Ошибка: {str(e)}",
                                 parent=list_window)
            return
        if not main_key_var.get():
            values = [str(row[0]).strip() for row in sheet.iter_rows(min_row=1, values_only=True) if
                      row and row[0] is not None]
            if not values:
                messagebox.showwarning("Ошибка", "В файле не найдено данных для импорта.", parent=list_window)
                return
            if messagebox.askyesno("Подтверждение импорта",
                                   f"Найдено {len(values)} значений для импорта.\n\n"
                                   "Это окно будет закрыто, и все текущие данные будут перезаписаны.\n"
                                   "Продолжить?",
                                   parent=list_window):
                sets_list[0]['widget_ref'].delete('1.0', END)
                sets_list[0]['widget_ref'].insert('1.0', '\n'.join(values))
                save_combobox()
        else:
            main_keys_data = {}
            current_main_key = None
            for i, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), 1):
                if len(row) < 3 or (row[1] is None or str(row[1]).strip() == "") or \
                        (row[2] is None or str(row[2]).strip() == ""):
                    if any(c is not None for c in row):
                        messagebox.showwarning("Ошибка",
                                               f"Строка {i}: Неверный формат. Ключ и значение должны быть заполнены.",
                                               parent=list_window)
                        return
                    continue
                main_key = str(row[0]).strip() if row[0] is not None else ""
                key = str(row[1]).strip()
                value = str(row[2]).strip()
                if main_key:
                    current_main_key = main_key
                    if current_main_key not in main_keys_data:
                        main_keys_data[current_main_key] = {}
                    main_keys_data[current_main_key][key] = value
                elif current_main_key:
                    if key in main_keys_data[current_main_key]:
                        messagebox.showwarning("Ошибка",
                                               f"Строка {i}: Дублирующийся ключ '{key}' для главного ключа '{current_main_key}'.",
                                               parent=list_window)
                        return
                    main_keys_data[current_main_key][key] = value
                else:
                    messagebox.showwarning("Ошибка",
                                           f"Строка {i}: Ключ '{key}' не имеет предшествующего главного ключа.",
                                           parent=list_window)
                    return
            if not main_keys_data:
                messagebox.showwarning("Ошибка", "В файле не найдено данных для импорта.", parent=list_window)
                return
            total_sub_keys = sum(len(d) for d in main_keys_data.values())
            if messagebox.askyesno("Подтверждение импорта",
                                   f"Найдено {len(main_keys_data)} главных ключей и {total_sub_keys} подчинённых ключей.\n\n"
                                   "Это окно будет закрыто, и все текущие данные будут перезаписаны.\n"
                                   "Продолжить?",
                                   parent=list_window):
                sets_list.clear()
                for main_key, sub_dict in main_keys_data.items():
                    new_set = {
                        "main_key": tk.StringVar(value=main_key),
                        "key_values": [{"key": tk.StringVar(value=k), "value": tk.StringVar(value=v)}
                                       for k, v in sub_dict.items()]
                    }
                    sets_list.append(new_set)
                save_combobox()

    # --- Window Layout ---
    top_controls_frame = tk.Frame(list_window)
    top_controls_frame.pack(fill="x", padx=10, pady=5)
    table_frame = tk.Frame(list_window)
    table_frame.pack(padx=10, pady=5, fill="both", expand=True)
    bottom_buttons_frame = tk.Frame(list_window)
    bottom_buttons_frame.pack(pady=10)

    tk.Label(top_controls_frame, text="Имя списка:").grid(row=0, column=0, sticky="w")
    tk.Entry(top_controls_frame, textvariable=name_var, width=40).grid(row=0, column=1, sticky="ew")
    if not is_edit:
        ttk.Checkbutton(top_controls_frame, text="Использовать главный ключ", variable=main_key_var,
                        command=refresh_table).grid(row=1, column=0, columnspan=2, pady=5)
    top_controls_frame.grid_columnconfigure(1, weight=1)

    tk.Button(bottom_buttons_frame, text="ОК", width=10, command=save_combobox).pack(side=LEFT, padx=5)
    tk.Button(bottom_buttons_frame, text="Отмена", width=10, command=list_window.destroy).pack(side=LEFT, padx=5)

    refresh_table()


# --- Combination Window Functions ---
def open_combination_window(listbox, item_to_edit, parent_window):
    """
    Window to create or edit a 'Combination' tag.
    If item_to_edit is provided, it opens in edit mode.
    """
    is_edit = item_to_edit is not None
    title = "Редактирование сочетания" if is_edit else "Создание сочетания"

    combo_window = tk.Toplevel(parent_window)
    combo_window.title(title)
    combo_window.resizable(False, False)
    combo_window.focus_set()
    combo_window.grab_set()

    name_var = tk.StringVar()
    combination_tags = []
    old_name = None

    if is_edit:
        old_name, _, _ = listbox.item(item_to_edit)['values']
        combo_config = load_json(COMBINATION_CONFIG_PATH, 'combination_config')
        existing_combo_data = next((item for item in combo_config if item['name'] == old_name), None)
        if existing_combo_data:
            name_var.set(existing_combo_data['name'])
            combination_tags.extend(existing_combo_data['tags'])
        else:
            messagebox.showwarning("Ошибка", "Данные сочетания не найдены.")
            return

    all_tags = get_all_tags_for_constructor()
    sorted_tags = sorted(list(all_tags))

    def refresh_listbox():
        for item in combo_listbox.get_children():
            combo_listbox.delete(item)
        for i, tag in enumerate(combination_tags):
            display_tag = tag
            if tag == ' ':
                display_tag = "[Пробел]"
            elif tag == '\n':
                display_tag = "[Абзац]"
            elif tag == 'today_tag':
                display_tag = "[СЕГОДНЯ]"
            combo_listbox.insert("", "end", values=(i + 1, display_tag))

    def add_element(element):
        if element:
            if element == "[Пробел]":
                element = ' '
            elif element == "[Абзац]":
                element = '\n'
            elif element == "[СЕГОДНЯ]":
                element = 'today_tag'

            combination_tags.append(element)
            refresh_listbox()

    def remove_element():
        selected_item = combo_listbox.selection()
        if selected_item:
            index_to_remove = int(combo_listbox.item(selected_item[0])['values'][0]) - 1
            if 0 <= index_to_remove < len(combination_tags):
                del combination_tags[index_to_remove]
                refresh_listbox()

    def move_element(direction):
        selected_item = combo_listbox.selection()
        if not selected_item:
            return
        index = int(combo_listbox.item(selected_item[0])['values'][0]) - 1
        if direction == 'up' and index > 0:
            combination_tags[index], combination_tags[index - 1] = combination_tags[index - 1], combination_tags[index]
        elif direction == 'down' and index < len(combination_tags) - 1:
            combination_tags[index], combination_tags[index + 1] = combination_tags[index + 1], combination_tags[index]
        refresh_listbox()
        new_selection_index = index + (1 if direction == 'down' else -1)
        combo_listbox.selection_set(combo_listbox.get_children()[new_selection_index])

    def save_combination():
        name = name_var.get().strip()
        if not name:
            messagebox.showwarning("Ошибка", "Введите имя для сочетания.", parent=combo_window)
            return
        if not combination_tags:
            messagebox.showwarning("Ошибка", "Добавьте хотя бы один элемент в сочетание.", parent=combo_window)
            return

        all_configs = (load_json(path, '') for path in
                       [FIELDS_CONFIG_PATH, COMBOBOX_REGULAR_PATH, COMBOBOX_MAINKEY_PATH, COMBINATION_CONFIG_PATH])
        all_names = {item['name'] for config in all_configs for item in config if item['name'] != old_name}
        if name in all_names:
            messagebox.showwarning("Ошибка", f"Имя '{name}' уже используется.", parent=combo_window)
            return

        combo_data = {"name": name, "type": "текст", "tag_type": "сочетание", "tags": combination_tags}
        combo_config = load_json(COMBINATION_CONFIG_PATH, 'combination_config')

        if is_edit:
            found = False
            for i, combo in enumerate(combo_config):
                if combo['name'] == old_name:
                    combo_config[i] = combo_data
                    found = True
                    break
            if not found:
                messagebox.showerror("Ошибка", "Не удалось найти сочетание для редактирования.")
                return
        else:
            combo_config.append(combo_data)

        save_json(COMBINATION_CONFIG_PATH, combo_config)
        refresh_all_windows(listbox)
        combo_window.destroy()

    content_frame = tk.Frame(combo_window, padx=10, pady=10)
    content_frame.pack(fill="both", expand=True)

    button_frame = tk.Frame(combo_window)
    button_frame.pack(side="bottom", pady=10)
    tk.Button(button_frame, text="ОК", width=10, command=save_combination).pack(side="left", padx=5)
    tk.Button(button_frame, text="Отмена", width=10, command=combo_window.destroy).pack(side="left", padx=5)

    tk.Label(content_frame, text="Имя сочетания:").pack(anchor="w")
    name_entry = tk.Entry(content_frame, textvariable=name_var, width=50)
    name_entry.pack(pady=(0, 10), fill="x")

    options_frame = tk.Frame(content_frame)
    options_frame.pack(fill="x")

    tag_options_frame = tk.Frame(options_frame)
    tag_options_frame.pack(side="left", padx=(0, 10))
    tk.Label(tag_options_frame, text="Добавить тег:").pack(anchor="w")
    tag_combo = ttk.Combobox(tag_options_frame, values=sorted_tags, state="readonly")
    tag_combo.pack(side="left", fill="x", expand=True, padx=(0, 5))
    tk.Button(tag_options_frame, text="Добавить", command=lambda: add_element(tag_combo.get())).pack(side="left")

    literal_options_frame = tk.Frame(options_frame)
    literal_options_frame.pack(side="left")
    tk.Label(literal_options_frame, text="Добавить текст:").pack(anchor="w")
    literal_entry = tk.Entry(literal_options_frame, width=20)
    literal_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))
    tk.Button(literal_options_frame, text="Добавить", command=lambda: add_element(literal_entry.get())).pack(
        side="left")

    literal_buttons_frame = tk.Frame(content_frame, pady=5)
    literal_buttons_frame.pack(fill="x")
    tk.Button(literal_buttons_frame, text="Пробел", command=lambda: add_element(' ')).pack(side="left", padx=2)
    tk.Button(literal_buttons_frame, text="Абзац", command=lambda: add_element('\n')).pack(side="left", padx=2)
    tk.Button(literal_buttons_frame, text="СЕГОДНЯ", command=lambda: add_element('today_tag')).pack(side="left", padx=2)

    tk.Label(content_frame, text="Элементы сочетания:").pack(anchor="w", pady=(10, 0))

    combo_listbox_frame = tk.Frame(content_frame)
    combo_listbox_frame.pack(fill="both", expand=True, pady=(0, 10))

    combo_listbox = ttk.Treeview(combo_listbox_frame, columns=("№", "Элемент"), show="headings", height=10)
    combo_listbox.heading("№", text="№", anchor="center")
    combo_listbox.heading("Элемент", text="Элемент")
    combo_listbox.column("№", width=30, anchor="center")
    combo_listbox.column("Элемент", width=400, anchor="w")

    listbox_scrollbar = ttk.Scrollbar(combo_listbox_frame, orient="vertical", command=combo_listbox.yview)
    combo_listbox.configure(yscrollcommand=listbox_scrollbar.set)
    combo_listbox.pack(side="left", fill="both", expand=True)
    listbox_scrollbar.pack(side="right", fill="y")

    actions_frame = tk.Frame(content_frame)
    actions_frame.pack(fill="x", pady=5)
    tk.Button(actions_frame, text="Удалить", width=12, command=remove_element).pack(side="left", padx=2)
    tk.Button(actions_frame, text="Вверх", width=12, command=lambda: move_element('up')).pack(side="left", padx=2)
    tk.Button(actions_frame, text="Вниз", width=12, command=lambda: move_element('down')).pack(side="left", padx=2)

    refresh_listbox()


def load_combination_config():
    """Load combination config from JSON file, initializing if not found."""
    try:
        with open(COMBINATION_CONFIG_PATH, encoding='utf8') as f:
            return json.load(f)
    except FileNotFoundError:
        with open(COMBINATION_CONFIG_PATH, 'w', encoding='utf8') as f:
            json.dump([], f, ensure_ascii=False, indent=4)
        messagebox.showwarning("Информация", f"Создан новый файл конфигурации: {COMBINATION_CONFIG_PATH}")
        return []
    except json.JSONDecodeError:
        messagebox.showwarning("Предупреждение",
                               f"Файл {COMBINATION_CONFIG_PATH} поврежден. Инициализация пустой конфигурации.")
        with open(COMBINATION_CONFIG_PATH, 'w', encoding='utf8') as f:
            json.dump([], f, ensure_ascii=False, indent=4)
        return []

def load_combination_config():
    """Load combination config from JSON file, initializing if not found."""
    try:
        with open(COMBINATION_CONFIG_PATH, encoding='utf8') as f:
            return json.load(f)
    except FileNotFoundError:
        with open(COMBINATION_CONFIG_PATH, 'w', encoding='utf8') as f:
            json.dump([], f, ensure_ascii=False, indent=4)
        messagebox.showwarning("Информация", f"Создан новый файл конфигурации: {COMBINATION_CONFIG_PATH}")
        return []
    except json.JSONDecodeError:
        messagebox.showwarning("Предупреждение", f"Файл {COMBINATION_CONFIG_PATH} поврежден. Инициализация пустой конфигурации.")
        with open(COMBINATION_CONFIG_PATH, 'w', encoding='utf8') as f:
            json.dump([], f, ensure_ascii=False, indent=4)
        return []


def open_edit_tag_window(listbox, parent_window):
    """Opens the appropriate window for editing a selected tag."""
    selected_item = listbox.selection()
    if not selected_item:
        messagebox.showwarning("Ошибка", "Выберите тег для редактирования.", parent=listbox)
        return

    item_id = selected_item[0]
    item_values = listbox.item(item_id)['values']

    if not item_values or len(item_values) < 3:
        messagebox.showwarning("Ошибка", "Некорректные данные тега.", parent=listbox)
        return

    tag_type = item_values[2]

    # Pass the listbox and the item_id to the specific edit window
    if tag_type == 'поле':
        open_field_window(listbox, item_id, parent_window)
    elif tag_type == 'чекбокс':
        open_checkbox_window(listbox, item_id, parent_window)
    elif tag_type == 'комбобокс' or tag_type == 'список':
        open_list_window(listbox, item_id, parent_window)
    elif tag_type == 'сочетание':
        open_combination_window(listbox, item_id, parent_window)
    else:
        messagebox.showinfo("Информация", "Редактирование для этого типа тега еще не реализовано.")


def delete_tag(tags_listbox, parent_window):
    """
    Deletes one or more selected tags and performs a cascading delete
    of their subkeys from rules and combinations.
    """
    selected_items = tags_listbox.selection()
    if not selected_items:
        messagebox.showwarning("Предупреждение", "Пожалуйста, выберите тег(и) для удаления.", parent=parent_window)
        return

    # Collect details of all selected tags for the confirmation message
    tags_to_delete = []
    for item_id in selected_items:
        item_values = tags_listbox.item(item_id)['values']
        tag_name = item_values[0]
        tag_type = item_values[2]
        tags_to_delete.append({'name': tag_name, 'type': tag_type})

    # Format a user-friendly confirmation message, now with a warning
    names_str = "\n- ".join([t['name'] for t in tags_to_delete])
    if not messagebox.askyesno("Подтверждение", f"Вы уверены, что хотите удалить следующие теги?\n\n- {names_str}\n\n"
                                               "ВНИМАНИЕ: Все выбранные теги и их дочерние элементы (для списков) будут также удалены из всех правил и сочетаний.",
                               parent=parent_window):
        return

    try:
        # --- START: New Cascading Delete Logic ---

        # 1. Collect all top-level tags to purge
        tags_to_purge = set(tag['name'] for tag in tags_to_delete)

        # 2. Find all subkeys from any 'список' type tags being deleted.
        subkeys_to_purge = set()
        list_tags_to_delete = [tag['name'] for tag in tags_to_delete if tag['type'] == 'список']

        if list_tags_to_delete:
            mainkey_config = load_json(COMBOBOX_MAINKEY_PATH, 'combobox_mainkey')
            for combo in mainkey_config:
                if combo.get('name') in list_tags_to_delete:
                    for mk_dict in combo.get('main_keys', []):
                        # mk_dict is like: {"Main Key 1": {"subkey1": "val1", "subkey2": "val2"}}
                        if mk_dict and isinstance(list(mk_dict.values())[0], dict):
                            subkeys_dict = list(mk_dict.values())[0]
                            for subkey_name in subkeys_dict.keys():
                                subkeys_to_purge.add(subkey_name)

        # 3. Add subkeys to the purge set
        tags_to_purge.update(subkeys_to_purge)

        # 4. If we have tags to purge, clean them up from other configs.
        if tags_to_purge:
            # 4a. Clean up Combination Config
            combination_config = load_json(COMBINATION_CONFIG_PATH, 'combination_config')
            for combo in combination_config:
                # Filter the 'tags' list, keeping only tags not in the purge set
                combo['tags'] = [tag for tag in combo.get('tags', []) if tag not in tags_to_purge]
            save_json(COMBINATION_CONFIG_PATH, combination_config)

            # 4b. Clean up Rules Config
            rules_config = load_json(RULES_CONFIG_PATH, 'rules_config')
            for rule in rules_config:
                # Filter 'conditions' by removing any that use a purged tag
                rule['conditions'] = [cond for cond in rule.get('conditions', []) if cond.get('tag') not in tags_to_purge]
                # Filter 'behaviors' similarly
                rule['behaviors'] = [beh for beh in rule.get('behaviors', []) if beh.get('tag') not in tags_to_purge]
            save_json(RULES_CONFIG_PATH, rules_config)

        # --- END: New Cascading Delete Logic ---


        # --- Original Deletion Logic (for the main tags themselves) ---
        deletions_by_file = {
            FIELDS_CONFIG_PATH: set(),
            COMBOBOX_REGULAR_PATH: set(),
            COMBOBOX_MAINKEY_PATH: set(),
            COMBINATION_CONFIG_PATH: set()
        }

        for tag in tags_to_delete:
            tag_name = tag['name']
            tag_type = tag['type']
            if tag_type in ('поле', 'чекбокс'):
                deletions_by_file[FIELDS_CONFIG_PATH].add(tag_name)
            elif tag_type == 'комбобокс':
                deletions_by_file[COMBOBOX_REGULAR_PATH].add(tag_name)
            elif tag_type == 'список':
                deletions_by_file[COMBOBOX_MAINKEY_PATH].add(tag_name)
            elif tag_type == 'сочетание':
                deletions_by_file[COMBINATION_CONFIG_PATH].add(tag_name)

        # Process each configuration file that has items to be deleted
        for config_path, names_to_delete_set in deletions_by_file.items():
            if not names_to_delete_set:
                continue

            config_data = load_json(config_path, '')
            updated_data = [item for item in config_data if item.get('name') not in names_to_delete_set]
            save_json(config_path, updated_data)

        # Refresh the UI once after all deletions are complete
        refresh_all_windows(tags_listbox)
        messagebox.showinfo("Успех", "Выбранные теги и все их дочерние элементы были успешно удалены.", parent=parent_window)

    except Exception as e:
        messagebox.showerror("Ошибка", f"Произошла ошибка при удалении тегов: {e}", parent=parent_window)


# This function is needed to refresh the constructor listbox from other windows
def populate_tags_listbox_in_constructor(listbox):
    for i in listbox.get_children():
        listbox.delete(i)

    all_items = []
    all_items.extend(
        [(f['name'], f['type'], f.get('tag_type', 'поле')) for f in load_json(FIELDS_CONFIG_PATH, 'fields_config')])
    all_items.extend([(c['name'], c['type'], c.get('tag_type', 'комбобокс')) for c in
                      load_json(COMBOBOX_REGULAR_PATH, 'combobox_regular')])
    all_items.extend([(c['name'], c['type'], c.get('tag_type', 'список')) for c in
                      load_json(COMBOBOX_MAINKEY_PATH, 'combobox_mainkey')])
    all_items.extend([(c['name'], c['type'], c.get('tag_type', 'сочетание')) for c in
                      load_json(COMBINATION_CONFIG_PATH, 'combination_config')])

    all_items.sort(key=lambda x: x[0])
    for item in all_items:
        listbox.insert("", "end", values=item)


def refresh_main_and_constructor():
    """Refreshes the dynamic widgets in the main window and the tags in the constructor window."""
    global dynamic_frame

    # 1. Save current state before destroying widgets
    current_state = get_current_widget_values(dynamic_frame)

    if dynamic_frame and dynamic_frame.winfo_exists():
        for widget in dynamic_frame.winfo_children():
            widget.destroy()
        # 2. Pass the saved state to the loading function
        load_all_dynamic_widgets(initial_state=current_state)

    # Find and refresh the constructor's tags listbox if it is open
    for child in window.winfo_children():
        if isinstance(child, tk.Toplevel) and child.title() == "Конструктор":
            for grand_child in child.winfo_children():
                if isinstance(grand_child, ttk.Notebook):
                    tags_tab = grand_child.winfo_children()[0]
                    for great_grand_child in tags_tab.winfo_children():
                        if isinstance(great_grand_child, tk.Frame):
                            for great_great_grand_child in great_grand_child.winfo_children():
                                if isinstance(great_great_grand_child, ttk.Treeview):
                                    populate_tags_listbox(great_great_grand_child)
                                    return

def evaluate_condition(condition, merge_data):
    """
    Evaluate a single condition against the main window's tag value.
    Args:
        condition (dict) {'tag': 'tag_name', 'condition': 'содержит', 'rule': 'value'}
        merge_data (dict): Data from get_common_merge_data()
    Returns:
        bool: True if condition is met, False otherwise
    """
    tag = condition['tag']
    cond_type = condition['condition']
    rule = condition['rule']

    if tag not in merge_data:
        return False  # Skip missing tags silently

    value = merge_data[tag]
    try:
        if cond_type == 'содержит':
            return rule.lower() in value.lower()
        elif cond_type == 'начинается с':
            return value.lower().startswith(rule.lower())
        elif cond_type == 'заканчивается на':
            return value.lower().endswith(rule.lower())
        elif cond_type == 'больше':
            return float(value) > float(rule)
        elif cond_type == 'меньше':
            return float(value) < float(rule)
        elif cond_type == 'равно':
            if tag in checkbox_vars:
                return value == rule
            else:
                return float(value) == float(rule)
        elif cond_type == 'True':
            return value == "1"  # Checkbox check
        elif cond_type == 'False':
            return value == "0"  # Checkbox check
        else:
            return False  # Unknown condition
    except ValueError:
        return False  # Invalid numeric conversion

def apply_behaviors(behaviors, merge_data):
    """
    Apply behaviors to merge_data.
    Args:
        behaviors (list): List of behavior dictionaries from rules_config.json
        merge_data (dict): Data from get_common_merge_data()
    Returns:
        dict: Modified merge_data
    """
    for behavior in behaviors:
        tag = behavior['tag']
        action = behavior['condition']
        rule = behavior['rule']
        if tag not in merge_data:
            continue
        value = str(merge_data[tag])  # Convert to string for safety
        try:
            if action == "очистить":
                merge_data[tag] = ""
            elif action == "очистить при не выполнении":
                merge_data[tag] = ""
            elif action == "CAPS":
                merge_data[tag] = value.upper()
            elif action == "верхняя буква":
                merge_data[tag] = value.capitalize()
            elif action == "нижняя буква":
                merge_data[tag] = value.lower()
            elif action == "транслит":
                merge_data[tag] = translit(value, rule)
            elif action == "добавить текст в начале":
                merge_data[tag] = rule + str(value)
            elif action == "добавить текст в конце":
                merge_data[tag] = str(value) + rule
            elif action == "добавить дату":
                date_obg = datetime.strptime(value, '%d.%m.%Y')
                days = int(rule)
                merge_data[tag] = (date_obg + timedelta(days=days)).strftime('%d.%m.%Y')
            elif action == "отнять дату":
                date_obg = datetime.strptime(value, '%d.%m.%Y')
                days = int(rule)
                merge_data[tag] = (date_obg - timedelta(days=days)).strftime('%d.%m.%Y')
            elif action == "обрезать":
                if ':' not in rule:
                    raise ValueError("Rule must be in 'start:end' format")
                start, end = map(int, rule.split(':'))
                if start <= end:
                    merge_data[tag] = value[:start] + value[end + 1:] if 0 <= start <= end < len(value) else value
                else:
                    merge_data[tag] = value[:end] + value[start + 1:] if 0 <= end <= start < len(value) else value
        except Exception as e:
            print(f"Error applying behavior {action} on {tag}: {e}")
    return merge_data


def open_create_rule_window(listbox, constructor_window):
    create_rule_window = tk.Toplevel(constructor_window)
    create_rule_window.title("Создать правило")
    create_rule_window.geometry("1070x230")
    create_rule_window.resizable(False, False)
    create_rule_window.focus_set()
    create_rule_window.grab_set()

    # Get all tags, sort them, and add a blank option at the start
    sorted_tags = sorted(list(set(get_all_tags_for_constructor())))
    sorted_tags.insert(0, '')

    main_frame = tk.Frame(create_rule_window)
    main_frame.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)

    widgets_3 = []
    widgets_4 = []
    rule_widgets = {}

    tk.Label(main_frame, text="Имя правила:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    name_var = tk.StringVar()
    tk.Entry(main_frame, textvariable=name_var, width=30).grid(row=0, column=1, columnspan=5, padx=5, pady=5, sticky="w")

    first_column_frame = tk.Frame(main_frame)
    first_column_frame.grid(row=2, column=0, columnspan=3, padx=5, pady=5, sticky="w")

    tk.Label(main_frame, text="Теги", anchor="center").grid(row=1, column=0, padx=5, pady=5)
    tk.Label(main_frame, text="Условие", anchor="center").grid(row=1, column=1, padx=5, pady=5)
    tk.Label(main_frame, text="Правило", anchor="center").grid(row=1, column=2, padx=5, pady=5)

    tag_var1 = tk.StringVar()
    tag_combobox1 = ttk.Combobox(first_column_frame, textvariable=tag_var1, width=20, state="readonly")
    tag_combobox1['values'] = sorted_tags
    tag_combobox1.grid(row=0, column=0, padx=5, pady=5)

    condition_var1 = tk.StringVar()
    condition_combobox1 = ttk.Combobox(first_column_frame, textvariable=condition_var1, width=20, state="readonly")
    # Add a blank option to the condition combobox
    condition_combobox1['values'] = ["", "содержит", "начинается с", "заканчивается на", "меньше", "больше", "равно", "True", "False"]
    condition_combobox1.grid(row=0, column=1, padx=5, pady=5)

    rule_entry_var1 = tk.StringVar()
    rule_entry1 = tk.Entry(first_column_frame, textvariable=rule_entry_var1, width=20)
    rule_entry1.grid(row=0, column=2, padx=5, pady=5)

    condition_combobox1.bind("<<ComboboxSelected>>",
                             lambda e: toggle_rule_entry_state(condition_combobox1, rule_entry1))

    widgets_3.append((first_column_frame, tag_combobox1, condition_combobox1, rule_entry1))

    def toggle_rule_entry_state(condition_combobox, rule_entry):
        selected_condition = condition_combobox.get()
        if selected_condition in ["True", "False"]:
            rule_entry.config(state="disabled")
        else:
            rule_entry.config(state="normal")

    def add_new_3_item_set():
        base_row = len(widgets_3)
        new_frame = tk.Frame(main_frame)
        new_frame.grid(row=2 + base_row, column=0, columnspan=3, padx=5, pady=5, sticky="w")

        new_tag_var = tk.StringVar()
        new_tag_combobox = ttk.Combobox(new_frame, textvariable=new_tag_var, width=20, state="readonly")
        new_tag_combobox['values'] = sorted_tags
        new_tag_combobox.grid(row=0, column=0, padx=5, pady=5)

        new_condition_var = tk.StringVar()
        new_condition_combobox = ttk.Combobox(new_frame, textvariable=new_condition_var, width=20, state="readonly")
        # Add a blank option to the new condition combobox
        new_condition_combobox['values'] = ["", "содержит", "начинается с", "заканчивается на", "меньше", "больше", "равно", "True", "False"]
        new_condition_combobox.grid(row=0, column=1, padx=5, pady=5)

        new_rule_var = tk.StringVar()
        new_rule_entry = tk.Entry(new_frame, textvariable=new_rule_var, width=20)
        new_rule_entry.grid(row=0, column=2, padx=5, pady=5)

        new_condition_combobox.bind("<<ComboboxSelected>>",
                                    lambda e: toggle_rule_entry_state(new_condition_combobox, new_rule_entry))

        widgets_3.append((new_frame, new_tag_combobox, new_condition_combobox, new_rule_entry))

        # Update button positions
        buttons_frame_3.grid(row=3 + base_row, column=0, columnspan=3, padx=5, pady=5, sticky="w")
        new_height = max(230, 230 + max(len(widgets_3), len(widgets_4)) * 40)
        create_rule_window.geometry(f"1070x{new_height}")

    def remove_last_3_item_set():
        if len(widgets_3) > 1:
            frame_to_destroy, _, _, _ = widgets_3.pop()
            frame_to_destroy.destroy()
            buttons_frame_3.grid(row=2 + len(widgets_3), column=0, columnspan=3, padx=5, pady=5, sticky="w")
            new_height = max(230, 230 + max(len(widgets_3), len(widgets_4)) * 40)
            create_rule_window.geometry(f"1070x{new_height}")

    buttons_frame_3 = tk.Frame(main_frame)
    buttons_frame_3.grid(row=2 + len(widgets_3), column=0, columnspan=3, padx=5, pady=5, sticky="w")
    ttk.Button(buttons_frame_3, text="Добавить", command=add_new_3_item_set).pack(side="left", padx=(0, 5))
    ttk.Button(buttons_frame_3, text="Удалить", command=remove_last_3_item_set).pack(side="left")

    ttk.Separator(main_frame, orient='vertical').grid(row=1, column=3, rowspan=100, sticky="ns", padx=10)

    second_column_frame = tk.Frame(main_frame)
    second_column_frame.grid(row=2, column=4, columnspan=4, padx=5, pady=5, sticky="w")

    tk.Label(main_frame, text="Теги", anchor="center").grid(row=1, column=4, padx=5, pady=5)
    tk.Label(main_frame, text="Условие", anchor="center").grid(row=1, column=5, padx=5, pady=5)
    tk.Label(main_frame, text="Правило", anchor="center").grid(row=1, column=6, padx=5, pady=5)
    tk.Label(main_frame, text="Опция", anchor="center").grid(row=1, column=7, padx=5, pady=5)

    tag_var2 = tk.StringVar()
    tag_combobox2 = ttk.Combobox(second_column_frame, textvariable=tag_var2, width=20, state="readonly")
    tag_combobox2['values'] = sorted_tags
    tag_combobox2.grid(row=0, column=0, padx=5, pady=5)

    condition_var2 = tk.StringVar()
    condition_combobox2 = ttk.Combobox(second_column_frame, textvariable=condition_var2, width=20, state="readonly")
    # Add a blank option to the condition combobox
    condition_combobox2['values'] = ["", "очистить", "очистить при не выполнении", "CAPS", "верхняя буква", "нижняя буква", "транслит",
                                     "добавить текст в начале", "добавить текст в конце", "добавить дату", "отнять дату", "обрезать"]
    condition_combobox2.grid(row=0, column=1, padx=5, pady=5)

    rule_entry_var2 = tk.StringVar()
    rule_entry2 = tk.Entry(second_column_frame, textvariable=rule_entry_var2, width=20)
    rule_entry2.grid(row=0, column=2, padx=5, pady=5)

    option_entry_var2 = tk.StringVar()
    option_entry2 = tk.Entry(second_column_frame, textvariable=option_entry_var2, width=20, state="disabled")
    option_entry2.grid(row=0, column=3, padx=5, pady=5)

    condition_combobox2.bind("<<ComboboxSelected>>",
                             lambda e: toggle_behavior_widgets(condition_combobox2, second_column_frame, option_entry2,
                                                               0, 2, rule_entry_var2))

    widgets_4.append((second_column_frame, tag_combobox2, condition_combobox2, rule_entry2, option_entry2))

    def toggle_behavior_widgets(condition_combobox, rule_frame, option_entry, index, col, rule_var):
        behavior = condition_combobox.get()
        frame, tag_cb, cond_cb, old_rule_widget, opt_entry = widgets_4[index]
        if old_rule_widget:
            old_rule_widget.destroy()
        if behavior == "транслит":
            rule_widget = ttk.Combobox(rule_frame, textvariable=rule_var, width=10, state="readonly")
            rule_widget['values'] = ["uk", "ru", "pl", "hu", "ro"]
            rule_widget.current(0)
            option_entry.config(state="disabled")
        else:
            rule_widget = tk.Entry(rule_frame, textvariable=rule_var, width=20)
            if behavior in ["очистить", "CAPS", "верхняя буква", "нижняя буква", "очистить при не выполнении"]:
                rule_widget.config(state="disabled")
                option_entry.config(state="disabled")
            else:
                rule_widget.config(state="normal")
                option_entry.config(state="disabled")
        rule_widget.grid(row=0, column=col, padx=5, pady=5)
        widgets_4[index] = (frame, tag_cb, cond_cb, rule_widget, opt_entry)

    def add_new_4_item_set():
        base_row = len(widgets_4)
        new_frame = tk.Frame(main_frame)
        new_frame.grid(row=2 + base_row, column=4, columnspan=4, padx=5, pady=5, sticky="w")

        new_tag_var = tk.StringVar()
        new_tag_combobox = ttk.Combobox(new_frame, textvariable=new_tag_var, width=20, state="readonly")
        new_tag_combobox['values'] = sorted_tags
        new_tag_combobox.grid(row=0, column=0, padx=5, pady=5)

        new_condition_var = tk.StringVar()
        new_condition_combobox = ttk.Combobox(new_frame, textvariable=new_condition_var, width=20, state="readonly")
        # Add a blank option to the new condition combobox
        new_condition_combobox['values'] = ["", "очистить", "очистить при не выполнении", "CAPS", "верхняя буква", "нижняя буква", "транслит",
                                            "добавить текст в начале", "добавить текст в конце", "добавить дату", "отнять дату", "обрезать"]
        new_condition_combobox.grid(row=0, column=1, padx=5, pady=5)

        new_option_var = tk.StringVar()
        new_option_entry = tk.Entry(new_frame, textvariable=new_option_var, width=20, state="disabled")
        new_option_entry.grid(row=0, column=3, padx=5, pady=5)

        new_rule_var = tk.StringVar()
        new_rule_entry = tk.Entry(new_frame, textvariable=new_rule_var, width=20)
        new_rule_entry.grid(row=0, column=2, padx=5, pady=5)

        new_condition_combobox.bind("<<ComboboxSelected>>",
                                    lambda e: toggle_behavior_widgets(new_condition_combobox, new_frame,
                                                                      new_option_entry, base_row, 2, new_rule_var))

        widgets_4.append((new_frame, new_tag_combobox, new_condition_combobox, new_rule_entry, new_option_entry))

        buttons_frame_4.grid(row=2 + len(widgets_4), column=4, columnspan=4, padx=5, pady=5, sticky="w")
        new_height = max(230, 230 + max(len(widgets_3), len(widgets_4)) * 40)
        create_rule_window.geometry(f"1070x{new_height}")

    def remove_last_4_item_set():
        if len(widgets_4) > 1:
            frame_to_destroy, _, _, _, _ = widgets_4.pop()
            frame_to_destroy.destroy()
            buttons_frame_4.grid(row=2 + len(widgets_4), column=4, columnspan=4, padx=5, pady=5, sticky="w")
            new_height = max(230, 230 + max(len(widgets_3), len(widgets_4)) * 40)
            create_rule_window.geometry(f"1070x{new_height}")

    buttons_frame_4 = tk.Frame(main_frame)
    buttons_frame_4.grid(row=2 + len(widgets_4), column=4, columnspan=4, padx=5, pady=5, sticky="w")
    ttk.Button(buttons_frame_4, text="Добавить", command=add_new_4_item_set).pack(side="left", padx=(0, 5))
    ttk.Button(buttons_frame_4, text="Удалить", command=remove_last_4_item_set).pack(side="left")

    def validate_and_save():
        rule_name = name_var.get().strip()
        if not rule_name:
            messagebox.showwarning("Ошибка", "Имя правила не может быть пустым.", parent=create_rule_window)
            return

        rules_config = load_json(RULES_CONFIG_PATH, 'rules_config')
        if rules_config is None:
            rules_config = []
            save_json(RULES_CONFIG_PATH, rules_config)
            messagebox.showwarning("Информация", f"Создан новый файл конфигурации: {RULES_CONFIG_PATH}", parent=create_rule_window)

        if any(rule['name'] == rule_name for rule in rules_config):
            messagebox.showwarning("Ошибка", "Имя правила уже существует.", parent=create_rule_window)
            return

        conditions = []
        for frame, tag_cb, cond_cb, rule_entry in widgets_3:
            tag = tag_cb.get()
            condition = cond_cb.get()
            rule = rule_entry.get() if isinstance(rule_entry, tk.Entry) else ""
            if tag and condition:
                if condition not in ["True", "False"] and not rule:
                    messagebox.showwarning("Ошибка", "Введите правило для условия.", parent=create_rule_window)
                    return
                if condition in ["меньше", "больше", "равно"]:
                    try:
                        float(rule)
                    except ValueError:
                        messagebox.showwarning("Ошибка",
                                               "Для условий 'меньше', 'больше' или 'равно' введите только число.", parent=create_rule_window)
                        return
                conditions.append({'tag': tag, 'condition': condition, 'rule': rule})

        behaviors = []
        for frame, tag_cb, cond_cb, rule_entry, opt_entry in widgets_4:
            tag = tag_cb.get()
            condition = cond_cb.get()
            rule = rule_entry.get() if isinstance(rule_entry, tk.Entry) else ""
            option = opt_entry.get()
            if tag and condition:
                if condition not in ["очистить", "CAPS", "верхняя буква", "нижняя буква",
                                     "очистить при не выполнении"] and not rule:
                    messagebox.showwarning("Ошибка", "Введите правило для поведения.")
                    return
                if condition in ["добавить дату", "отнять дату"] and not rule.isdigit():
                    messagebox.showwarning("Ошибка",
                                           "Для 'добавить дату' или 'отнять дату' введите количество дней в виде числа.", parent=create_rule_window)
                    return
                if condition == "обрезать" and not all(part.isdigit() for part in rule.split(':')):
                    messagebox.showwarning("Ошибка", "Для 'обрезать' введите диапазон в формате 'start:end' с числами.", parent=create_rule_window)
                    return
                behaviors.append({'tag': tag, 'condition': condition, 'rule': rule})

        if not conditions and not behaviors:
            messagebox.showwarning("Ошибка", "Правило должно содержать хотя бы одно условие или поведение.", parent=create_rule_window)
            return

        new_rule = {
            'name': rule_name,
            'conditions': conditions,
            'behaviors': behaviors
        }

        rules_config.append(new_rule)
        save_json(RULES_CONFIG_PATH, rules_config)
        messagebox.showinfo("Успех", f"Правило '{rule_name}' успешно создано.", parent=create_rule_window)
        update_rules_listbox(rules_config, listbox)
        listbox.update_idletasks()
        create_rule_window.destroy()

    button_frame = tk.Frame(create_rule_window)
    button_frame.pack(pady=20, fill=tk.X)
    button_frame.columnconfigure(0, weight=1)
    button_frame.columnconfigure(1, weight=1)
    tk.Button(button_frame, text="ОК", width=10, command=validate_and_save).grid(row=0, column=0, padx=5)
    tk.Button(button_frame, text="ОТМЕНА", width=10, command=create_rule_window.destroy).grid(row=0, column=1, padx=5)


def open_edit_rule_window(listbox, constructor_window):
    selected_item = listbox.selection()
    if not selected_item:
        messagebox.showwarning("Ошибка", "Выберите правило для редактирования", parent=constructor_window)
        return

    item_data = listbox.item(selected_item)
    rule_name = item_data['values'][0]

    rules_config = load_json(RULES_CONFIG_PATH, 'rules_config')
    rule_to_edit = next((rule for rule in rules_config if rule['name'] == rule_name), None)

    if not rule_to_edit:
        messagebox.showerror("Ошибка", f"Правило '{rule_name}' не найдено.")
        return

    edit_rule_window = tk.Toplevel(constructor_window)
    edit_rule_window.title(f"Изменить правило: {rule_name}")
    edit_rule_window.geometry("1070x230")
    edit_rule_window.resizable(False, False)
    edit_rule_window.focus_set()
    edit_rule_window.grab_set()

    # Store StringVars to prevent garbage collection
    edit_rule_window._vars_to_keep = []

    # Get all tags including subkeys, sort them, and add a blank option
    sorted_tags = sorted(list(set(get_all_tags_for_constructor())))
    sorted_tags.insert(0, '')

    main_frame = tk.Frame(edit_rule_window)
    main_frame.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)

    widgets_3 = []
    widgets_4 = []

    tk.Label(main_frame, text="Имя правила:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    name_var = tk.StringVar(value=rule_name)
    tk.Entry(main_frame, textvariable=name_var, width=30).grid(row=0, column=1, columnspan=5, padx=5, pady=5, sticky="w")

    first_column_frame = tk.Frame(main_frame)
    first_column_frame.grid(row=2, column=0, columnspan=3, padx=5, pady=5, sticky="w")

    tk.Label(main_frame, text="Теги", anchor="center").grid(row=1, column=0, padx=5, pady=5)
    tk.Label(main_frame, text="Условие", anchor="center").grid(row=1, column=1, padx=5, pady=5)
    tk.Label(main_frame, text="Правило", anchor="center").grid(row=1, column=2, padx=5, pady=5)

    buttons_frame_3 = tk.Frame(main_frame)
    buttons_frame_4 = tk.Frame(main_frame)

    def toggle_rule_entry_state(condition_combobox, rule_entry):
        selected_condition = condition_combobox.get()
        if selected_condition in ["True", "False"]:
            rule_entry.config(state="disabled")
        else:
            rule_entry.config(state="normal")

    def add_new_3_item_set(pre_populated_data=None):
        base_row = len(widgets_3)
        new_frame = tk.Frame(main_frame)
        new_frame.grid(row=2 + base_row, column=0, columnspan=3, padx=5, pady=5, sticky="w")

        new_tag_var = tk.StringVar()
        new_tag_combobox = ttk.Combobox(new_frame, textvariable=new_tag_var, width=20, state="readonly")
        new_tag_combobox['values'] = sorted_tags
        new_tag_combobox.grid(row=0, column=0, padx=5, pady=5)

        new_condition_var = tk.StringVar()
        new_condition_combobox = ttk.Combobox(new_frame, textvariable=new_condition_var, width=20, state="readonly")
        new_condition_combobox['values'] = [
            "", "содержит", "начинается с", "заканчивается на", "меньше", "больше", "равно", "True", "False"
        ]
        new_condition_combobox.grid(row=0, column=1, padx=5, pady=5)

        new_rule_var = tk.StringVar()
        new_rule_entry = tk.Entry(new_frame, textvariable=new_rule_var, width=20)
        new_rule_entry.grid(row=0, column=2, padx=5, pady=5)

        widgets_3.append((new_frame, new_tag_combobox, new_condition_combobox, new_rule_entry))
        edit_rule_window._vars_to_keep.extend([new_tag_var, new_condition_var, new_rule_var])

        new_condition_combobox.bind("<<ComboboxSelected>>",
                                    lambda e: toggle_rule_entry_state(new_condition_combobox, new_rule_entry))

        if pre_populated_data:
            new_tag_var.set(pre_populated_data.get('tag', ''))
            new_condition_var.set(pre_populated_data.get('condition', ''))
            new_rule_var.set(pre_populated_data.get('rule', ''))
            toggle_rule_entry_state(new_condition_combobox, new_rule_entry)

        buttons_frame_3.grid(row=2 + len(widgets_3), column=0, columnspan=3, padx=5, pady=5, sticky="w")
        new_height = max(230, 230 + max(len(widgets_3), len(widgets_4)) * 40)
        edit_rule_window.geometry(f"1070x{new_height}")

    def remove_last_3_item_set():
        if len(widgets_3) > 1:
            frame_to_destroy, _, _, _ = widgets_3.pop()
            frame_to_destroy.destroy()
            buttons_frame_3.grid(row=2 + len(widgets_3), column=0, columnspan=3, padx=5, pady=5, sticky="w")
            new_height = max(230, 230 + max(len(widgets_3), len(widgets_4)) * 40)
            edit_rule_window.geometry(f"1070x{new_height}")

    if rule_to_edit.get('conditions'):
        for condition in rule_to_edit.get('conditions'):
            add_new_3_item_set(condition)
    else:
        add_new_3_item_set()

    buttons_frame_3.grid(row=2 + len(widgets_3), column=0, columnspan=3, padx=5, pady=5, sticky="w")
    ttk.Button(buttons_frame_3, text="Добавить", command=add_new_3_item_set).pack(side="left", padx=(0, 5))
    ttk.Button(buttons_frame_3, text="Удалить", command=remove_last_3_item_set).pack(side="left")

    ttk.Separator(main_frame, orient='vertical').grid(row=1, column=3, rowspan=100, sticky="ns", padx=10)

    second_column_frame = tk.Frame(main_frame)
    second_column_frame.grid(row=2, column=4, columnspan=4, padx=5, pady=5, sticky="w")

    tk.Label(main_frame, text="Теги", anchor="center").grid(row=1, column=4, padx=5, pady=5)
    tk.Label(main_frame, text="Условие", anchor="center").grid(row=1, column=5, padx=5, pady=5)
    tk.Label(main_frame, text="Правило", anchor="center").grid(row=1, column=6, padx=5, pady=5)
    tk.Label(main_frame, text="Опция", anchor="center").grid(row=1, column=7, padx=5, pady=5)

    def toggle_behavior_widgets(condition_combobox, rule_frame, option_entry, index, col, rule_var):
        behavior = condition_combobox.get()
        frame, tag_cb, cond_cb, old_rule_widget, opt_entry = widgets_4[index]
        if old_rule_widget:
            old_rule_widget.destroy()
        if behavior == "транслит":
            rule_widget = ttk.Combobox(rule_frame, textvariable=rule_var, width=10, state="readonly")
            rule_widget['values'] = ["uk", "ru", "pl", "hu", "ro"]
            rule_widget.current(0)
            option_entry.config(state="disabled")
        else:
            rule_widget = tk.Entry(rule_frame, textvariable=rule_var, width=20)
            if behavior in ["очистить", "CAPS", "верхняя буква", "нижняя буква", "очистить при не выполнении"]:
                rule_widget.config(state="disabled")
                option_entry.config(state="disabled")
            else:
                rule_widget.config(state="normal")
                option_entry.config(state="disabled")
        rule_widget.grid(row=0, column=col, padx=5, pady=5)
        widgets_4[index] = (frame, tag_cb, cond_cb, rule_widget, opt_entry)

    def add_new_4_item_set(pre_populated_data=None):
        base_row = len(widgets_4)
        new_frame = tk.Frame(main_frame)
        new_frame.grid(row=2 + base_row, column=4, columnspan=4, padx=5, pady=5, sticky="w")

        new_tag_var = tk.StringVar()
        new_tag_combobox = ttk.Combobox(new_frame, textvariable=new_tag_var, width=20, state="readonly")
        new_tag_combobox['values'] = sorted_tags
        new_tag_combobox.grid(row=0, column=0, padx=5, pady=5)

        new_condition_var = tk.StringVar()
        new_condition_combobox = ttk.Combobox(new_frame, textvariable=new_condition_var, width=20, state="readonly")
        new_condition_combobox['values'] = [
            "", "очистить", "очистить при не выполнении", "CAPS", "верхняя буква", "нижняя буква", "транслит",
            "добавить текст в начале", "добавить текст в конце", "добавить дату", "отнять дату", "обрезать"
        ]
        new_condition_combobox.grid(row=0, column=1, padx=5, pady=5)

        new_option_var = tk.StringVar()
        new_option_entry = tk.Entry(new_frame, textvariable=new_option_var, width=20, state="disabled")
        new_option_entry.grid(row=0, column=3, padx=5, pady=5)

        new_rule_var = tk.StringVar()
        new_rule_entry = tk.Entry(new_frame, textvariable=new_rule_var, width=20)
        new_rule_entry.grid(row=0, column=2, padx=5, pady=5)

        widgets_4.append((new_frame, new_tag_combobox, new_condition_combobox, new_rule_entry, new_option_entry))
        edit_rule_window._vars_to_keep.extend([new_tag_var, new_condition_var, new_rule_var, new_option_var])

        new_condition_combobox.bind("<<ComboboxSelected>>",
                                    lambda e: toggle_behavior_widgets(new_condition_combobox, new_frame,
                                                                      new_option_entry, base_row, 2, new_rule_var))

        if pre_populated_data:
            new_tag_var.set(pre_populated_data.get('tag', ''))
            new_condition_var.set(pre_populated_data.get('condition', ''))
            new_rule_var.set(pre_populated_data.get('rule', ''))
            new_option_var.set(pre_populated_data.get('option', ''))
            toggle_behavior_widgets(new_condition_combobox, new_frame, new_option_entry, base_row, 2, new_rule_var)

        buttons_frame_4.grid(row=2 + len(widgets_4), column=4, columnspan=4, padx=5, pady=5, sticky="w")
        new_height = max(230, 230 + max(len(widgets_3), len(widgets_4)) * 40)
        edit_rule_window.geometry(f"1070x{new_height}")

    def remove_last_4_item_set():
        if len(widgets_4) > 1:
            frame_to_destroy, _, _, _, _ = widgets_4.pop()
            frame_to_destroy.destroy()
            buttons_frame_4.grid(row=2 + len(widgets_4), column=4, columnspan=4, padx=5, pady=5, sticky="w")
            new_height = max(230, 230 + max(len(widgets_3), len(widgets_4)) * 40)
            edit_rule_window.geometry(f"1070x{new_height}")

    if rule_to_edit.get('behaviors'):
        for behavior in rule_to_edit.get('behaviors'):
            add_new_4_item_set(behavior)
    else:
        add_new_4_item_set()

    buttons_frame_4.grid(row=2 + len(widgets_4), column=4, columnspan=4, padx=5, pady=5, sticky="w")
    ttk.Button(buttons_frame_4, text="Добавить", command=add_new_4_item_set).pack(side="left", padx=(0, 5))
    ttk.Button(buttons_frame_4, text="Удалить", command=remove_last_4_item_set).pack(side="left")

    def validate_and_save_edit():
        new_rule_name = name_var.get().strip()
        if not new_rule_name:
            messagebox.showwarning("Ошибка", "Имя правила не может быть пустым.", parent=edit_rule_window)
            return

        rules_config = load_json(RULES_CONFIG_PATH, 'rules_config')
        if rules_config is None: rules_config = []

        # Check for name duplication, excluding the current rule being edited
        if new_rule_name != rule_name and any(rule['name'] == new_rule_name for rule in rules_config):
            messagebox.showwarning("Ошибка", "Имя правила уже существует.", parent=edit_rule_window)
            return

        conditions = []
        for frame, tag_cb, cond_cb, rule_entry in widgets_3:
            tag = tag_cb.get()
            condition = cond_cb.get()
            rule = rule_entry.get()
            if tag and condition:
                if condition not in ["True", "False"] and not rule:
                    messagebox.showwarning("Ошибка", "Введите правило для условия.", parent=edit_rule_window)
                    return
                if condition in ["меньше", "больше", "равно"]:
                    try:
                        float(rule)
                    except ValueError:
                        messagebox.showwarning("Ошибка",
                                               "Для условий 'меньше', 'больше' или 'равно' введите только число.", parent=edit_rule_window)
                        return
                conditions.append({'tag': tag, 'condition': condition, 'rule': rule})

        behaviors = []
        for frame, tag_cb, cond_cb, rule_entry, opt_entry in widgets_4:
            tag = tag_cb.get()
            condition = cond_cb.get()
            rule = rule_entry.get()
            option = opt_entry.get()
            if tag and condition:
                if condition not in ["очистить", "CAPS", "верхняя буква", "нижняя буква",
                                     "очистить при не выполнении"] and not rule:
                    messagebox.showwarning("Ошибка", "Введите правило для поведения.")
                    return
                if condition in ["добавить дату", "отнять дату"] and not rule.isdigit():
                    messagebox.showwarning("Ошибка",
                                           "Для 'добавить дату' или 'отнять дату' введите количество дней в виде числа.", parent=edit_rule_window)
                    return
                if condition == "обрезать" and not all(part.isdigit() for part in rule.split(':')):
                    messagebox.showwarning("Ошибка", "Для 'обрезать' введите диапазон в формате 'start:end' с числами.", parent=edit_rule_window)
                    return
                behaviors.append({'tag': tag, 'condition': condition, 'rule': rule})

        updated_rules = [rule for rule in rules_config if rule['name'] != rule_name]
        updated_rules.append({
            'name': new_rule_name,
            'conditions': conditions,
            'behaviors': behaviors
        })

        save_json(RULES_CONFIG_PATH, updated_rules)
        messagebox.showinfo("Успех", f"Правило '{new_rule_name}' успешно обновлено.", parent=edit_rule_window)
        update_rules_listbox(updated_rules, listbox)
        edit_rule_window.destroy()

    button_frame = tk.Frame(edit_rule_window)
    button_frame.pack(pady=20, fill=tk.X)
    button_frame.columnconfigure(0, weight=1)
    button_frame.columnconfigure(1, weight=1)
    tk.Button(button_frame, text="ОК", width=10, command=validate_and_save_edit).grid(row=0, column=0, padx=5)
    tk.Button(button_frame, text="ОТМЕНА", width=10, command=edit_rule_window.destroy).grid(row=0, column=1, padx=5)


def delete_rule(listbox, parent_window, rules_file=RULES_CONFIG_PATH):
    """Deletes one or more selected rules."""
    selected_items = listbox.selection()
    if not selected_items:
        messagebox.showwarning("Ошибка", "Выберите правило(а) для удаления", parent=parent_window)
        return

    # Collect the names of all selected rules
    rules_to_delete_names = [listbox.item(item_id)['values'][0] for item_id in selected_items]

    # Format a user-friendly confirmation message
    names_str = "\n- ".join(rules_to_delete_names)
    response = messagebox.askyesno("Подтверждение",
                                   f"Вы уверены, что хотите удалить следующие правила?\n\n- {names_str}",
                                   parent=parent_window)
    if not response:
        return

    try:
        existing_rules = load_json(rules_file, 'rules_config')
        if not isinstance(existing_rules, list):  # Basic validation
            existing_rules = []

        # Use a set for efficient filtering
        rules_to_delete_set = set(rules_to_delete_names)
        updated_rules = [rule for rule in existing_rules if rule.get('name') not in rules_to_delete_set]

        save_json(rules_file, updated_rules)

        # Update the listbox with the new data
        update_rules_listbox(updated_rules, listbox)
        messagebox.showinfo("Успех", "Выбранные правила были успешно удалены.", parent=parent_window)

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить изменения: {str(e)}", parent=parent_window)


# --- Main Window Creation ---
# Create a single Tkinter window instance
window = tk.Tk()
window.title("Doxy v2.1")
window.resizable(False, False)

# Determine the base directory
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Path to icon
ICON_DIR = os.path.join(BASE_DIR, 'icon')
ICON_PATH = os.path.join(ICON_DIR, "bee.ico")

# Set main window icon
try:
    window.iconbitmap(ICON_PATH)
except tk.TclError:
    print("Warning: Icon file not found or invalid format.")

# Apply icon automatically to all Toplevel windows
_original_toplevel_init = tk.Toplevel.__init__
def _custom_toplevel_init(self, *args, **kwargs):
    _original_toplevel_init(self, *args, **kwargs)
    try:
        self.iconbitmap(ICON_PATH)
    except tk.TclError:
        pass

tk.Toplevel.__init__ = _custom_toplevel_init

# Main layout frames
# This frame will hold the dynamic widgets and expand with them
dynamic_frame = tk.Frame(window, padx=10, pady=10)
dynamic_frame.pack(fill="both", expand=True)

# A separator for visual clarity
ttk.Separator(window, orient='horizontal').pack(fill='x', pady=5)

# This frame holds the static buttons at the bottom
bottom_frame = tk.Frame(window)
bottom_frame.pack(fill="x", padx=10, pady=(0, 10))

# --- Bottom Buttons ---
constructor_button = ttk.Button(bottom_frame, text="Конструктор", command=open_constructor_window)
report_button = ttk.Button(bottom_frame, text="Сформировать", command=submit_and_save)
ttk.Button(bottom_frame, text="Импорт", command=import_fields).pack(side=LEFT, padx=5, pady=5, fill=X, expand=True)
constructor_button.pack(side=LEFT, padx=5, pady=5, fill=X, expand=True)
report_button.pack(side=LEFT, padx=5, pady=5, fill=X, expand=True)

# --- Initial Load and Main Loop ---
window.bind_all("<Key>", _onKeyRelease, "+")
load_all_dynamic_widgets()
window.mainloop()