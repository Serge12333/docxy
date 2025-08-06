import tkinter
from tkinter import *
from tkinter import messagebox
from tkinter.ttk import *
import os
from mailmerge import MailMerge
from transliterate import translit
import tkentrycomplete
import json
import shutil
import openpyxl
from datetime import datetime, timedelta
from decimal import Decimal
import uuid


def _onKeyRelease(event):
    ctrl = (event.state & 0x4) != 0
    if event.keycode == 88 and ctrl and event.keysym.lower() != "x":
        event.widget.event_generate("<<Cut>>")
    elif event.keycode == 86 and ctrl and event.keysym.lower() != "v":
        event.widget.event_generate("<<Paste>>")
    elif event.keycode == 67 and ctrl and event.keysym.lower() != "c":
        event.widget.event_generate("<<Copy>>")


def load_json(file_path, dict_name):
    try:
        with open(file_path, encoding='utf8') as f:
            return json.load(f)
    except Exception as ex:
        print(ex)
        return {}


materialDict = load_json('D:\\document_filler\\materialDict.json', 'materialDict')
materialDict1 = load_json('D:\\document_filler\\materialDict1.json', 'materialDict1')


def selected(func):
    data = materialDict.get(var_material.get(), {})
    name_companys.set(data.get("name_company", ""))
    street_companys.set(data.get("street_company", ""))
    index_companys.set(data.get("index_company", ""))
    town_companys.set(data.get("town_company", ""))
    town_company_ukrs.set(data.get("town_company_ukr", ""))
    country_companys.set(data.get("country_company", ""))
    country_company_ukrs.set(data.get("country_company_ukr", ""))
    isos.set(data.get("iso", ""))


def selected1(func):
    data = materialDict1.get(var_material1.get(), {})
    name_company_dels.set(data.get("name_company_del", ""))
    street_company_dels.set(data.get("street_company_del", ""))
    index_company_dels.set(data.get("index_company_del", ""))
    town_company_dels.set(data.get("town_company_del", ""))
    town_company_del_ukrs.set(data.get("town_company_del_ukr", ""))
    country_company_dels.set(data.get("country_company_del", ""))
    country_company_ukr_dels.set(data.get("country_company_ukr_del", ""))


def format_decimal(value, entry):
    try:
        # Convert to Decimal directly for precise calculation
        decimal_value = Decimal(str(value).replace(',', '.'))
        final_result = decimal_value.quantize(Decimal('0.00'))
        formatted = '{:,.2f}'.format(final_result).replace(',', ' ').replace('.', ',')
        return formatted
    except (ValueError, Decimal.InvalidOperation):
        messagebox.showwarning(title="Предупреждение!",
                               message="Допускается использование только чисел с \"точкой\" или \"запятой\" в качестве десятичного разделителя")
        return "0,00"


def summ():
    try:
        # Replace comma with dot for float conversion and handle user input flexibility
        price = float(pricee.get().replace(',', '.'))
        netto_weight = float(netto_weighte.get().replace(',', '.'))
        result = format_decimal(price * netto_weight, pricee)
        summ_label.config(text=result)
        return result
    except ValueError as e:
        messagebox.showwarning(title="Предупреждение!",
                               message="Введите корректные числовые значения для цены и веса нетто")
        summ_label.config(text="0,00")
        return "0,00"


def brutto_weight():
    try:
        tare_weight = float(tare_weighte.get().replace(',', '.'))
        netto_weight = float(netto_weighte.get().replace(',', '.'))
        result = format_decimal(tare_weight + netto_weight, tare_weighte)
        brutto_weight_label.config(text=result)
        return result
    except ValueError as e:
        messagebox.showwarning(title="Предупреждение!",
                               message="Введите корректные числовые значения для веса тары и веса нетто")
        brutto_weight_label.config(text="0,00")
        return "0,00"


def get_common_merge_data():
    final_pricee2 = format_decimal(float(pricee.get().replace(',', '.')), pricee)
    final_tare_weighte2 = format_decimal(float(tare_weighte.get().replace(',', '.')), tare_weighte)
    final_netto_weighte2 = format_decimal(float(netto_weighte.get().replace(',', '.')), netto_weighte)

    # Construct addres_of_delivery with label and name_company_dels
    delivery_data = materialDict1.get(var_material1.get(), {})
    addres_of_delivery = f"Адреса доставки/Delivery address: {delivery_data.get('name_company_del', '')}: " \
                         f"{delivery_data.get('street_company_del', '')}, {delivery_data.get('index_company_del', '')} " \
                         f"{delivery_data.get('town_company_del', '')}, {delivery_data.get('country_company_del', '')}".strip()

    return {
        'cved_number': cved_numbere.get(),
        'cved_date': cved_datee.get(),
        'driver': drivere.get(),
        'eta_border': eta_bordere.get(),
        'license_plate': license_platee.get(),
        'drums_quantity': drums_quantitye.get(),
        'number_batch': number_batche.get(),
        'netto_weight': final_netto_weighte2,
        'cmr_number': cmr_numbere.get(),
        'cmr_numberd': cmr_numberde.get(),
        'seal_number': seal_numbere.get(),
        'invoice_number': invoice_numbere.get(),
        'vetinq_date': vetinq_datee.get(),
        'contract_number': contract_numbere.get(),
        'contract_date': contract_datee.get(),
        'prod_date': prod_datee.get(),
        'exp_concl': exp_concle.get(),
        'exp_concld': exp_conclde.get(),
        'decl_prod': decl_prode.get(),
        'decl_prodd': decl_prodde.get(),
        'invoice_date': invoice_datee.get(),
        'natprod_num': natprod_nume.get(),
        'cved_number_ukr': translit(cved_numbere.get(), "uk").upper(),
        'price': final_pricee2,
        'tare_weight': final_tare_weighte2,
        'payment': paymentc.get(),
        'incoterms': incotermsc.get().upper(),
        'license_platez': license_platee.get()[:8],
        'license_trailer': license_platee.get()[9:],
        'name_company': name_companys.get(),
        'street_company': street_companys.get(),
        'index_company': index_companys.get(),
        'town_company': town_companys.get(),
        'town_company_ukr': town_company_ukrs.get(),
        'country_company': country_companys.get(),
        'country_company_ukr': country_company_ukrs.get(),
        'iso': isos.get(),
        'crop': crope.get(),
        'name_company_del': name_company_dels.get(),
        'street_company_del': street_company_dels.get(),
        'index_company_del': index_company_dels.get(),
        'town_company_del': town_company_dels.get(),
        'town_company_del_ukr': town_company_del_ukrs.get(),
        'country_company_del': country_company_dels.get(),
        'country_company_ukr_del': country_company_ukr_dels.get(),
        'summ': summ(),
        'brutto_weight': brutto_weight(),
        'addres_of_delivery': addres_of_delivery,
        'delivery_costs_eng': "",
        'delivery_costs_rus': ""
    }


def merge_huzar_specifics(merge_data):
    print("Merging Huzar specifics...")  # Debug print
    material = var_material.get() or ""
    print(f"Material value: '{material}'")  # Debug material value with quotes
    if material == "Huzar":
        print("Huzar selected, updating all sunflower values...")  # Debug
        merge_data.update({
            'sunflowerrus': ", соняшниковий",
            'sunflowerukr': ", соняшниковий",
            'sunflowerpol': ", słonecznik",
            'sunflowereng': ", sunflower"
        })
    else:
        print("Huzar not selected, setting all sunflower fields to blank...")  # Debug
        merge_data.update({
            'sunflowerrus': "",
            'sunflowerukr': "",
            'sunflowereng': "",
            'sunflowerpol': ""
        })
    print(f"Updated merge_data keys and values: {list(merge_data.items())}")  # Debug all key-value pairs
    return merge_data


def merge_invoice_specifics(merge_data, checkpoints_val, checkpoints_ukr_val):
    material = var_material.get() or ""
    bank_value = banksc.get() or ""
    merge_data['bankv'] = bank_value

    # Initialize merge text with defaults
    merge_text = {
        'address_fur': "",
        'delivery_costs_rus': "",
        'delivery_costs_eng': ""
    }

    # Populate based on material
    ukr_del_value = ukr_del.get() or ""
    delivery_costs = delivery_costse.get() or ""
    town_ukr = town_company_del_ukrs.get() or ""
    town = town_company_dels.get() or ""

    if material == "Breitsamer":
        address_text = f"Загальна сума включає транспортні витрати від Польського кордону до {town_ukr} € {delivery_costs} \n Total amount includes transport cost from Polish border to {town} € {delivery_costs}"
        merge_text.update({
            'address_fur': address_text
        })
    elif material in ["FURSTEN-REFORM", "Langnese"]:
        address_text = f"Загальна сума включає в себе транспортні витрати / Total amount includes transport costs: \n" \
                       f"1. від Запоріжжя, Україна до границі ЄС {checkpoints_ukr_val or ''} {ukr_del_value} EUR / " \
                       f"from Zaporizhya, Ukraine to the EU border {checkpoints_val or ''} is {ukr_del_value} EUR.\n" \
                       f"2. від границі ЄС {checkpoints_ukr_val or ''} до {town_ukr}, {delivery_costs} EUR / " \
                       f"from the EU border {checkpoints_val or ''} to {town} is {delivery_costs} EUR."
        merge_text.update({
            'address_fur': address_text
        })

    print(f"Merge text for {material}: {merge_text}")  # Debug
    merge_data.update(merge_text)

    return merge_data


def submit_and_save():
    try:
        dateinv = cmr_numberde.get()
        date_format = '%d.%m.%Y'
        dtObj = datetime.strptime(dateinv, date_format)
        cmrd_minus_one_str = (dtObj - timedelta(days=1)).strftime(date_format)

        source_dir = "D:\\document_filler\\template"  # Fixed path
        output_dir = os.path.join(source_dir, f"{invoice_numbere.get()}_{var_material1.get()}")

        # Validation checks before folder creation
        if materialDict[var_material.get()]["name_company"] != name_companys.get():
            messagebox.showwarning(title="Предупреждение!", message="Заполните поле \"покупатель\"")
            return False
        if materialDict1[var_material1.get()]["name_company_del"] != name_company_dels.get():
            messagebox.showwarning(title="Предупреждение!", message="Заполните поле \"грузополучатель\"")
            return False
        if var_material.get() in ["Breitsamer", "Langnese", "FURSTEN-REFORM", "Huzar"] and delivery_costse.get() == "":
            messagebox.showwarning(title="Предупреждение!", message="Заполните поле \"стоим доставки\"")
            return False
        if var_material.get() in ["Breitsamer", "Langnese", "FURSTEN-REFORM", "Huzar"] and ukr_del.get() == "":
            messagebox.showwarning(title="Предупреждение!", message="Заполните поле \"стоим доставки укр\"")
            return False

        # Create folder only if all validations pass
        if not os.path.exists(output_dir):
            os.mkdir(output_dir)
            templates = {
                'cved_DOR': '.cved_DOR.docx',
                'cved_KOR': '.cved_KOR.docx',
                'cved_ZAH': '.cved_ZAH.docx',
                'vet_inq_f2': '.vet_inq_f2.docx',
                'inq_rend_serv': '.inq_rend_serv.docx',
                'vet_derhspozh': '.vet_derhspozh.docx',
                'natur_honey_f2': '.natur_honey_f2.docx',
                'invoice': '.invoice.docx',
                'raw_note_border': '.raw_note_border.docx',
                'raw_note_customs': '.raw_note_customs.docx',
                'CMR_pattern': '.CMR_pattern.docx',
                'health_DOR': '.health_DOR.docx',
                'health_KOR': '.health_KOR.docx',
                'health_ZAH': '.health_ZAH.docx',
                'health_SIR': '.health_SIR.docx'
            }

            merges = {key: MailMerge(os.path.join(source_dir, template)) for key, template in templates.items()}
            common_data = get_common_merge_data()

            # Checkpoint logic
            border_map = {
                "Краківець - Корчова, Польща": ("(Korczowa, Poland)", "(Корчова, Польща)"),
                "Ягодин - Дорохуск, Польща": ("(Doroshusk, Poland)", "(Дорохуск, Польща)"),
                "Чоп - Захонь, Угорщина": ("(Zahony, Hungary)", "(Захонь, Угорщина)"),
                "Порубне - Сірет, Румунія": ("(Siret, Romania)", "(Сірет, Румунія)")
            }
            checkpoints, checkpoints_ukr = border_map.get(optmenu.get(), ("", ""))

            # Merge common data for all documents
            for merge in merges.values():
                merge.merge(**common_data)

            # Apply Huzar specifics to specified templates
            huzar_data = merge_huzar_specifics(common_data.copy())
            for key in ['invoice', 'natur_honey_f2', 'cved_ZAH', 'cved_KOR', 'cved_DOR', 'CMR_pattern']:
                try:
                    merges[key].merge(**huzar_data)
                except Exception:
                    pass  # Silently handle errors

            # Specific merges from your version
            for key in ['cved_DOR', 'cved_KOR', 'cved_ZAH', 'natur_honey_f2', 'CMR_pattern']:
                merges[key].merge(**merge_huzar_specifics(common_data.copy()))

            merges['invoice'].merge(**merge_invoice_specifics(common_data.copy(), checkpoints, checkpoints_ukr))

            merges['vet_inq_f2'].merge(
                border=optmenu.get(),
                languagecert=language_certe.get()
            )

            merges['inq_rend_serv'].merge(
                cmr_date_minusone=cmrd_minus_one_str
            )

            merges['vet_derhspozh'].merge(
                cmr_date_minusone=cmrd_minus_one_str,
                border=optmenu.get(),
                languagecert=language_certe.get(),
                languagecert_add=", " + language_certe_add.get()
            )

            merges['raw_note_border'].merge(
                weight_exp=weight_expe.get()
            )

            if var_material1.get() == "DHI (Eystrup)":
                merges['CMR_pattern'].merge(
                    name_company_del="",
                    street_company_del=""
                )

            # File operations
            border_files = {
                "Краківець - Корчова, Польща": [(merges['cved_KOR'], "CVED_KOR.docx"), (merges['health_KOR'], "health_KOR.docx")],
                "Ягодин - Дорохуск, Польща": [(merges['cved_DOR'], "CVED_DOR.docx"), (merges['health_DOR'], "health_DOR.docx")],
                "Чоп - Захонь, Угорщина": [(merges['cved_ZAH'], "CVED_ZAH.docx"), (merges['health_ZAH'], "health_ZAH.docx")],
                "Порубне - Сірет, Румунія": [(merges['cved_ZAH'], "CVED_ZAH.docx"), (merges['health_SIR'], "health_SIR.docx")]
            }
            default_files = [
                (merges['cved_DOR'], "CVED_DOR.docx"),
                (merges['cved_KOR'], "CVED_KOR.docx"),
                (merges['cved_ZAH'], "CVED_ZAH.docx"),
                (merges['health_DOR'], "health_DOR.docx"),
                (merges['health_KOR'], "health_KOR.docx"),
                (merges['health_ZAH'], "health_ZAH.docx"),
                (merges['health_SIR'], "health_SIR.docx")
            ]

            for merge, filename in border_files.get(borders.get(), default_files):
                merge.write(os.path.join(output_dir, f"{invoice_numbere.get()}_{filename}"))

            merges['vet_inq_f2'].write(os.path.join(output_dir, f"{invoice_numbere.get()}_запит на вет.св-ва Ф2.docx"))
            merges['inq_rend_serv'].write(os.path.join(output_dir, f"{invoice_numbere.get()}_ЗАЯВА про надання послуг 2021.docx"))
            merges['vet_derhspozh'].write(os.path.join(output_dir, f"{invoice_numbere.get()}_Заявка Держпродспроживслужба.docx"))
            merges['natur_honey_f2'].write(os.path.join(output_dir, f"{invoice_numbere.get()}_Натур мед- ФР фор №2-новая.docx"))
            merges['invoice'].write(os.path.join(output_dir, f"{invoice_numbere.get()}_Invoice.docx"))
            merges['raw_note_border'].write(os.path.join(output_dir, f"{invoice_numbere.get()}_сырьевая справка ВЕТ-ПОГР.docx"))
            merges['raw_note_customs'].write(os.path.join(output_dir, f"{invoice_numbere.get()}_сырьевая справка там.docx"))
            merges['CMR_pattern'].write(os.path.join(output_dir, f"{invoice_numbere.get()}_шаблон СМР.docx"))

            # Excel operations
            wb = openpyxl.load_workbook(os.path.join(source_dir, ".Packing_list.xlsx"))
            wb['Sheet1'].title = f"{invoice_numbere.get()}_{var_material1.get()}"
            sheet = wb.active
            sheet.cell(row=2, column=1, value=name_companys.get())
            sheet.cell(row=8, column=3, value=f"№{invoice_numbere.get()} from {invoice_datee.get()}")
            sheet.cell(row=7, column=9, value=prod_datee.get())
            sheet.cell(row=8, column=9, value=license_platee.get())
            sheet.cell(row=9, column=9, value=f"{contract_numbere.get()} dated {contract_datee.get()}")
            sheet.cell(row=7, column=13, value=number_batche.get())
            sheet.cell(row=38, column=7, value=drums_quantitye.get())
            sheet.cell(row=40, column=8, value=invoice_datee.get())
            sheet.cell(row=4, column=1, value=f"ПАКУВАЛЬНИЙ ЛИСТ №{invoice_numbere.get()}/ PACKING LIST №{invoice_numbere.get()}")
            sheet.cell(row=9, column=3, value="Honey natural homogenized, sunflower" if var_material.get() == "Huzar" else "Natural homogenized honey")
            wb.save(os.path.join(output_dir, f"{invoice_numbere.get()} Packing_list.xlsx"))

            for merge in merges.values():
                merge.close()

            messagebox.showinfo(title="Успех!", message=f"Набор документов {invoice_numbere.get()}_{var_material1.get()}.docx был успешно сформирован")
        else:
            messagebox.showwarning(title="Предупреждение!", message="Удалите либо переместите предыдущий набор документов")
            return False

    except ValueError:
        messagebox.showwarning(title="Предупреждение!", message="Заполните поля \"сумма\" и \"вес брутто\" или введите корректный формат даты CMR (д.м.г)")
        return False
    except PermissionError:
        pass


# GUI Setup
window = Tk()
window.resizable(width=False, height=False)
window.geometry('520x605')
window.title("document_Filler, v. 1.50")
window.bind_all("<Key>", _onKeyRelease, "+")
icon = PhotoImage(file="D:\\document_filler\\bee.png")
window.iconphoto(True, icon)

# Variables
var_material = StringVar()
var_material1 = StringVar()
borders = StringVar()
name_companys = StringVar()
street_companys = StringVar()
index_companys = StringVar()
town_companys = StringVar()
town_company_ukrs = StringVar()
country_companys = StringVar()
country_company_ukrs = StringVar()
isos = StringVar()
name_company_dels = StringVar()
street_company_dels = StringVar()
index_company_dels = StringVar()
town_company_dels = StringVar()
town_company_del_ukrs = StringVar()
country_company_dels = StringVar()
country_company_ukr_dels = StringVar()

# Frames
veterinary = Frame(window)
shipping = Frame(window)
sale = Frame(window)
production = Frame(window)
sale.grid(row=1, column=0)
production.grid(row=1, column=2)
shipping.grid(row=3, column=0)
veterinary.grid(row=3, column=2)

# Comboboxes
combo_material = tkentrycomplete.Combobox(sale, values=list(materialDict.keys()), justify="center",
                                          textvariable=var_material, width=17)
combo_material.set_completion_list(materialDict)
for event in ['<<ComboboxSelected>>', '<Right>', '<Return>', '<Leave>', '<Motion>', '<Button-1>', '<Button-2>']:
    combo_material.bind(event, func=selected)
combo_material.grid(row=0, column=1)

combo_material1 = tkentrycomplete.Combobox(shipping, values=list(materialDict1.keys()), justify="center",
                                           textvariable=var_material1, width=17)
combo_material1.set_completion_list(materialDict1)
for event in ['<<ComboboxSelected>>', '<Right>', '<Return>', '<Leave>', '<Motion>', '<Button-1>', '<Button-2>']:
    combo_material1.bind(event, func=selected1)
combo_material1.grid(row=1, column=1)

# Entries
cved_numbere = Entry(veterinary)
cved_datee = Entry(veterinary)
drivere = Entry(shipping)
eta_bordere = Entry(shipping)
license_platee = Entry(sale)
drums_quantitye = Entry(production)
number_batche = Entry(production)
netto_weighte = Entry(production)
cmr_numbere = Entry(shipping)
cmr_numberde = Entry(shipping)
seal_numbere = Entry(production)
invoice_numbere = Entry(sale)
vetinq_datee = Entry(veterinary)
contract_numbere = Entry(sale)
contract_datee = Entry(sale)
prod_datee = Entry(production)
exp_concle = Entry(veterinary)
exp_conclde = Entry(veterinary)
decl_prode = Entry(production)
decl_prodde = Entry(production)
invoice_datee = Entry(sale)
natprod_nume = Entry(veterinary)
pricee = Entry(sale)
tare_weighte = Entry(production)
weight_expe = Entry(veterinary)
delivery_costse = Entry(shipping)
ukr_del = Entry(shipping)
crope = Combobox(production, values=["2021", "2022", "2023", "2024", "2025", "2026"], width=17)
crope.current(1)
doctore = Entry(veterinary)
doctore_hun = Entry(veterinary)
doctore_pol = Entry(veterinary)
time_loadinge = Entry(veterinary)
doctor_translite = Entry(veterinary)
language_certe = Combobox(veterinary, values=["польська", "угорська", "румунська"], width=17)
language_certe.current(0)
banksc = Combobox(sale, values=[
    "Bank account UA133808050000000026005563335\nBENEFICIARY: NATURAL HONEY LLC, 40022, st. Psilska 19, office 41, Sumy, Ukraine\nBENEFICIARY BANK: Raiffeisen Bank JSC,\nGenerala Almazova street 4A, KYIV 01011, Ukraine\nSWIFTcode: AVALUAUK\nCorrespondent Account: 55.022.305\nCORRESPONDENT BANK: Raiffeisen Bank International AG\nSWIFTcode: RZBAATWW",
    "Intermediary Bank: Commerzbank AG, Frankfurt am Main, Germany\nSWIFT code - COBA DE FF\nCorrespondent Bank: FIRST UKRAINIAN INTERNATIONAL BANK, Kiev, Ukraine\nSWIFT code - FUIB UA 2X\nBeneficiary’s Bank: MOTOR BANK, JSC, Zaporozhye, Ukraine\nAccount: UA 813348510000016006804559119\nSWIFT code – MOOJUA2Z\nBeneficiary’s name, address: NATURAL HONEY LLC, 40022, Sumy, st. Psilska 19, office 41\nIBAN Account No UA413130090000026006001018612"
], width=17)
language_certe_add = Combobox(veterinary, values=["німецька", "французька", "бельгійська", "чеська", "румунська"],
                              width=17)
language_certe_add.current(0)
optmenu = Combobox(shipping,
                   values=["Краківець - Корчова, Польща", "Ягодин - Дорохуск, Польща", "Чоп - Захонь, Угорщина",
                           "Порубне - Сірет, Румунія"], width=17, textvariable=borders)
optmenu.current(0)
paymentc = Combobox(sale, values=[
    "100% prepayment",
    "50% of the payment 2-3 days before dispatching \nfrom the seller`s warehouse,\n 50% of the payment in 14 days\nafter arrival of the goods",
    "80% within 2 day after arrival of the Goods\nand the rest 20% within 21 days\nfrom the date of the delivery",
    "100% in 14 days after arrival",
    "14 days after arrival and approval",
    "15 days post-payment",
    "20 days after delivery and quality control of the goods",
    "21 days after arrival and approval",
    "50% at the moment of arrival of the goods and 50% within 30 days after the receipt of the goods",
    "100% 30 days after arrival and approval"
], width=17)
incotermsc = Combobox(shipping, values=["DAP", "FCA Zaporizhia", "EXW Zaporizhia"], width=17, justify="center")
incotermsc.current(0)

# Labels
veterinaryl = Label(window, text="Ветеринарка", foreground="grey")
shippingl = Label(window, text="Отгрузка", foreground="grey")
salel = Label(window, text="Продажа", foreground="grey")
productionl = Label(window, text="Производство", foreground="grey")
summ_label = Label(sale, text="0,00")
brutto_weight_label = Label(production, text="0,00")
paymentl = Label(sale, text="оплата:")
cved_numberl = Label(veterinary, text="номер CVED/здоровья:")
cved_datel = Label(veterinary, text="дата CVED/здоровья:")
driverl = Label(shipping, text="водитель:")
eta_borderl = Label(shipping, text="время прибытия:")
license_platel = Label(sale, text="номер машины:")
drums_quantityl = Label(production, text="количество бочек:")
number_batchl = Label(production, text="номер партии:")
netto_weightl = Label(production, text="вес нетто:")
cmr_numberl = Label(shipping, text="номер CMR:")
cmr_numberdl = Label(shipping, text="дата CMR:")
seal_numberl = Label(production, text="номер пломб:")
invoice_numberl = Label(sale, text="номер инвойса:")
vetinq_datel = Label(veterinary, text="дата вет свидотства (Ф2):")
contract_numberl = Label(sale, text="номер контракта:")
contract_datel = Label(sale, text="дата контракта:")
prod_datel = Label(production, text="дата производства:")
exp_concll = Label(veterinary, text="номер экспретного:")
exp_concldl = Label(veterinary, text="дата экспертного:")
decl_prodl = Label(production, text="деклар производителя:")
decl_proddl = Label(production, text="дата декл производител:")
invoice_datel = Label(sale, text="дата инвойса:")
natprod_numl = Label(veterinary, text="номер держспожив:")
pricel = Label(sale, text="цена за кг:")
name_comboboxl = Label(sale, text="покупатель:")
tare_weightl = Label(production, text="вес тары:")
summl = Label(sale, text="сумма:")
brutto_weightl = Label(production, text="вес брутто:")
optmenul = Label(shipping, text="граница:")
consigneel = Label(shipping, text="грузополучатель:")
incotermsl = Label(shipping, text="incoterms:")
language_certl = Label(veterinary, text="язык сертификата:")
language_certl_add = Label(veterinary, text="язык сертификата (доп):")
weight_expl = Label(veterinary, text="вес экспертный:")
delivery_costsl = Label(shipping, text="стоим. доставки:")
ukr_dell = Label(shipping, text="стоим. доставки укр:")
cropl = Label(production, text="урожай, год:")
doctorl = Label(veterinary, text="имя врача:")
doctorl_pol = Label(veterinary, text="имя врача (Пол):")
doctorl_hun = Label(veterinary, text="имя врача (Вен):")
time_loadingl = Label(veterinary, text="время загрузки:")
doctor_translitl = Label(veterinary, text="имя врача (транслит):")
bankl = Label(sale, text="банк:")

# Buttons
summ_button = Button(sale, text="Посчитать", command=summ)
brutto_button = Button(production, text="Посчитать", command=brutto_weight)
report_button = Button(window, text="Сформировать", command=submit_and_save)

# Grid Layout
paymentc.grid(row=1, column=1)
cved_numbere.grid(row=0, column=1)
cved_datee.grid(row=1, column=1)
drivere.grid(row=5, column=1)
eta_bordere.grid(row=6, column=1)
license_platee.grid(row=9, column=1)
drums_quantitye.grid(row=7, column=1)
number_batche.grid(row=5, column=1)
netto_weighte.grid(row=1, column=1)
cmr_numbere.grid(row=3, column=1)
cmr_numberde.grid(row=4, column=1)
seal_numbere.grid(row=0, column=1)
invoice_numbere.grid(row=5, column=1)
vetinq_datee.grid(row=6, column=1)
language_certe.grid(row=7, column=1)
language_certe_add.grid(row=8, column=1)
contract_numbere.grid(row=7, column=1)
contract_datee.grid(row=8, column=1)
prod_datee.grid(row=6, column=1)
exp_concle.grid(row=3, column=1)
exp_conclde.grid(row=4, column=1)
decl_prode.grid(row=8, column=1)
decl_prodde.grid(row=9, column=1)
invoice_datee.grid(row=6, column=1)
natprod_nume.grid(row=5, column=1)
pricee.grid(row=2, column=1)
tare_weighte.grid(row=2, column=1)
weight_expe.grid(row=9, column=1)
doctore.grid(row=10, column=1)
doctore_hun.grid(row=11, column=1)
doctore_pol.grid(row=12, column=1)
delivery_costse.grid(row=7, column=1)
ukr_del.grid(row=8, column=1)
crope.grid(row=10, column=1)
time_loadinge.grid(row=13, column=1)
doctor_translite.grid(row=14, column=1)

paymentl.grid(row=1, column=0)
cved_numberl.grid(row=0, column=0)
cved_datel.grid(row=1, column=0)
driverl.grid(row=5, column=0)
eta_borderl.grid(row=6, column=0)
license_platel.grid(row=9, column=0)
drums_quantityl.grid(row=7, column=0)
number_batchl.grid(row=5, column=0)
netto_weightl.grid(row=1, column=0)
cmr_numberl.grid(row=3, column=0)
cmr_numberdl.grid(row=4, column=0)
seal_numberl.grid(row=0, column=0)
invoice_numberl.grid(row=5, column=0)
vetinq_datel.grid(row=6, column=0)
contract_numberl.grid(row=7, column=0)
contract_datel.grid(row=8, column=0)
prod_datel.grid(row=6, column=0)
exp_concll.grid(row=3, column=0)
exp_concldl.grid(row=4, column=0)
decl_prodl.grid(row=8, column=0)
decl_proddl.grid(row=9, column=0)
invoice_datel.grid(row=6, column=0)
natprod_numl.grid(row=5, column=0)
pricel.grid(row=2, column=0)
tare_weightl.grid(row=2, column=0)
summl.grid(row=3, column=0)
brutto_weightl.grid(row=3, column=0)
optmenul.grid(row=0, column=0)
consigneel.grid(row=1, column=0)
incotermsl.grid(row=2, column=0)
language_certl.grid(row=7, column=0)
language_certl_add.grid(row=8, column=0)
weight_expl.grid(row=9, column=0)
doctorl.grid(row=10, column=0)
doctorl_hun.grid(row=11, column=0)
doctorl_pol.grid(row=12, column=0)
delivery_costsl.grid(row=7, column=0)
ukr_dell.grid(row=8, column=0)
cropl.grid(row=10, column=0)
time_loadingl.grid(row=13, column=0)
doctor_translitl.grid(row=14, column=0)
name_comboboxl.grid(row=0, column=0)

combo_material1.grid(row=1, column=1)
optmenu.grid(row=0, column=1)
incotermsc.grid(row=2, column=1)

summ_label.grid(row=3, column=1)
brutto_weight_label.grid(row=3, column=1)
summ_button.grid(row=4, column=1)
brutto_button.grid(row=4, column=1)
report_button.place(x=210, y=575)
veterinaryl.grid(row=2, column=2)
shippingl.grid(row=2, column=0)
salel.grid(row=0, column=0)
productionl.grid(row=0, column=2)

window.mainloop()